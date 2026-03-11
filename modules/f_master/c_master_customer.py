from datetime import datetime
import json
from fastapi import HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from library import *
import os
from library.router import app
from library.db import Db
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook


class c_master_customer(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        is_pusat,
        filter_other="",
        filter_other_conj="",
    ):
        if orderby == None or orderby == "":
            orderby = "id_customer ASC"

        filter_other = f"A.company_id = '{company_id}' AND A.cabang_id = '{cabang_id}'"
        filter_other_conj = "and"
        if company_id == 1 and cabang_id == 1:
            filter_other = f""
            filter_other_conj = f""
        elif company_id != 1 and is_pusat == True:
            filter_other = f"A.company_id = '{company_id}'"
            filter_other_conj = "and"

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            """
    SELECT A.*, b.company_name, C.cabang_name from master_customer A
LEFT JOIN master_company B on A.company_id = B.id_company
LEFT JOIN master_company_cabang C on A.company_id = C.id_company AND A.cabang_id = C.id_cabang
"""
            + str_clause
        )
        sql_count = (
            """
    SELECT COUNT(*) from master_customer A
LEFT JOIN master_company B on A.company_id = B.id_company
LEFT JOIN master_company_cabang C on A.company_id = C.id_company AND A.cabang_id = C.id_cabang
"""
            + str_clause_count
        )
        print(sql)

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data

    async def export_customer(
        self,
        company_id,
        cabang_id,
        is_pusat,
    ):

        filter_other = (
            f"WHERE A.company_id = '{company_id}' AND A.cabang_id = '{cabang_id}'"
        )

        if company_id == 1 and cabang_id == 1:
            filter_other = f""
            filter_other_conj = f""
        elif company_id != 1 and is_pusat == True:
            filter_other = f"WHERE A.company_id = '{company_id}'"

        sql = f"""
            SELECT A.*, b.company_name, C.cabang_name from master_customer A
            LEFT JOIN master_company B on A.company_id = B.id_company
            LEFT JOIN master_company_cabang C on A.company_id = C.id_company AND A.cabang_id = C.id_cabang 
            {filter_other}
                    """

        result = await self.db.executeToDict(sql)
        print(sql)

        wb = self.excel_return(result)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def excel_return(self, result_data):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "ID SO"
        ws["B1"].value = "ID Invoice"
        ws["C1"].value = "Account VA"
        ws["D1"].value = "Nama Customer"
        ws["E1"].value = "Alamat"
        ws["F1"].value = "NPWP"
        ws["G1"].value = "KTP"
        ws["H1"].value = "Nama Produk"
        ws["I1"].value = "Tanggal"
        ws["J1"].value = "Nama Company"
        ws["K1"].value = "Nama Cabang"
        ws["L1"].value = "Quantity"
        ws["M1"].value = "Harga Satuan"
        ws["N1"].value = "Harga Total"
        ws["O1"].value = "PPN %"
        ws["P1"].value = "PPN Value"
        ws["Q1"].value = "PPH 22 %"
        ws["R1"].value = "PPH 22 Value"
        ws["S1"].value = "Biaya Admin"
        ws["T1"].value = "Grand Total"
        ws["U1"].value = "Pembayaran"
        ws["V1"].value = "SalesMan"
        ws["W1"].value = "Status Pembayaran"
        ws["X1"].value = "Sisa Pembayaran"
        ws["Y1"].value = "Harga Satuan HPP"
        ws["Z1"].value = "Harga Total HPP"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            # print(key, value)
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key[:-5]:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb

    async def create(self, data):

        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
            }
        )

        sqlString = self.db.genStrInsertSingleObject(data, "master_customer")

        try:
            # print(sqlString)
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            print(e)
            message = {"status": "error : " + str(e)}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def update(self, data, data_where):

        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
            }
        )

        sqlString = self.db.genUpdateObject(data, data_where, "master_customer")
        # print(sqlString)
        try:
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def delete(self, data_where):
        sqlString = self.db.genDeleteObject(data_where, "master_customer")
        try:
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def get_customer(self):
        sql = f"""SELECT id_customer as value,nama_customer as text 
				    FROM master_customer 
                    WHERE status_release = 't' AND status_aktif = 't'
                    ORDER BY id_customer ASC
                    LIMIT 100"""
        result = await self.db.executeToDict(sql)
        # print(result)
        return result

    async def get_customer_where_condition(self, where_condition):
        if where_condition != None:
            where_sql = f"""WHERE {where_condition['where_condition']} AND status_release = 't' AND status_aktif = 't'"""
        else:
            where_sql = f"""WHERE (1=1)"""

        sql = f"""SELECT 
            id_customer AS value, 
            CONCAT(nama_customer, ' (', LEFT(alamat, 30), ')') AS text,
            is_pph
        FROM 
            master_customer 
        {where_sql}
        ORDER BY 
            id_customer ASC 
        LIMIT 20;"""

        print(sql)
        result = await self.db.executeToDict(sql)
        # print(result)
        return result

    async def get_atribut_customer(self, id_customer):
        sql = f"""SELECT id_customer as value,nama_customer as text,* FROM master_customer 
                  WHERE id_customer = '{id_customer}' AND status_release = 't' AND status_aktif = 't'
                  LIMIT 1"""
        result = await self.db.executeToDict(sql)
        data = {"data": result}

        # print(sql)
        return data


"""
list your path url at bottom
example /testing url
test from postman :
url/api/c_master_customer/testing
for post method and other method, check tutorial from 
https://fastapi.tiangolo.com/
"""


@app.get("/api/f_master/c_master_customer/read")
async def read_data(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    # print("the data:", nik, limit, orderby, offset, filter)
    ob_data = c_master_customer()
    return await ob_data.read(
        orderby, limit, offset, filter, company_id, cabang_id, is_pusat
    )


@app.post("/api/f_master/c_master_customer/create")
async def create_data(request: Request):
    data = await request.json()
    ob_data = c_master_customer()
    return await ob_data.create(data)


@app.post("/api/f_master/c_master_customer/update")
async def update_data(request: Request):
    data = await request.json()
    ob_data = c_master_customer()
    return await ob_data.update(data["update_data"], data["update_where"])


@app.post("/api/f_master/c_master_customer/delete")
async def delete(request: Request):
    data = await request.json()
    ob_data = c_master_customer()
    return await ob_data.delete(data)


@app.get("/api/f_master/c_master_customer/get_customer")
async def get_customer():
    ob_data = c_master_customer()
    return await ob_data.get_customer()


@app.get("/api/f_master/c_master_customer/get_customer_where_condition")
async def get_customer_where_condition(param: object = Query(None, alias="param")):
    print(param)
    data_where = json.loads(param)
    ob_data = c_master_customer()
    return await ob_data.get_customer_where_condition(data_where)


@app.get("/api/f_master/c_master_customer/get_atribut_customer")
async def get_atribut_customer(param: object = Query(None, alias="param")):
    # print('MASUKKKKKK')
    data = json.loads(param)
    ob_data = c_master_customer()
    return await ob_data.get_atribut_customer(data)
