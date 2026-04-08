from datetime import datetime
import io
import json
from fastapi import HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from library import *
import os
from library.router import app
from library.db import Db
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_dashboard_utama(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def get_inventori(
        self, company_id, cabang_id, tanggal, tanggal_start, is_pusat
    ):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year

        where = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND tanggal BETWEEN '2025-01-01' AND '{tanggal}' and stock_condition = 'good'"""

        if int(company_id) == 1:
            where = f"""WHERE tanggal BETWEEN '2025-01-01' AND '{tanggal}'  and stock_condition = 'good'"""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE company_id = {company_id} AND tanggal BETWEEN '2025-01-01' AND '{tanggal}'  and stock_condition = 'good'"""

        sql = f"""SELECT sum(ht) as qty
            FROM
            (
                SELECT
                company_id,
                cabang_id,
            SUM ( CASE WHEN in_out = 'IN' THEN harga_total ELSE 0 END ) - SUM ( CASE WHEN in_out = 'OUT' THEN harga_total ELSE 0 END ) ht 
            FROM
            trans_inventory_detail_mutasi 
            {where}
            GROUP BY
            produk_id,
            company_id,
            cabang_id 
            ) aa;"""

        try:
            result = await self.db.executeToDict(sql)
            # print(result)
            if len(result) == 0:
                return 0
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return result[0]["qty"]

    async def get_sales(self, company_id, cabang_id, tanggal, tanggal_start, is_pusat):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year

        where = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND tanggal_invoice BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        if int(company_id) == 1:
            where = f"""WHERE aa.tanggal_invoice BETWEEN '{tanggal_start}' AND '{tanggal}'"""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE company_id = {company_id} AND aa.tanggal_invoice BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        sql = f"""SELECT SUM (AA.amount) AS harga_total 
                FROM
                trans_inventory_subsidiary_invoice aa
                LEFT JOIN trans_inventory_subsidiary_sales_order_header bb ON bb.id_trans = aa.id_trans_sales_order
                {where}
            """

        # print(sql)
        try:
            result = await self.db.executeToDict(sql)
            print("\n\n\n\n", result)
            if len(result) == 0:
                return 0
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return result[0]["harga_total"]

    async def get_outstanding_payment(
        self, company_id, cabang_id, tanggal, tanggal_start, is_pusat
    ):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year

        where = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        if int(company_id) == 1:
            where = (
                f"""WHERE tanggal_invoice BETWEEN '{tanggal_start}' AND '{tanggal}'"""
            )
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE company_id = {company_id} AND tanggal_invoice BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        sql = f"""
                SELECT SUM
                ( amount_total_outstanding ) AS outstanding_payment 
                FROM
                trans_inventory_subsidiary_invoice aa
                LEFT JOIN trans_inventory_subsidiary_sales_order_header bb ON bb.id_trans = aa.id_trans_sales_order
                {where}
                """
        print(sql)

        try:
            result = await self.db.executeToDict(sql)
            # print(result)
            if len(result) == 0:
                return 0
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))
        print("\n\n\n\n")
        return result[0]["outstanding_payment"]

    async def get_paid_payment(
        self, company_id, cabang_id, tanggal, tanggal_start, is_pusat
    ):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year

        where = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""
        gb = f"""GROUP BY company_id,cabang_id"""

        if int(company_id) == 1:
            where = f"""WHERE payment_last_updated BETWEEN '{tanggal_start}' AND '{tanggal}'"""
            gb = ""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE company_id = {company_id} AND tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""
            gb = "GROUP BY company_id"

        # sql = f"""SELECT
        #             company_id,
        #             cabang_id,
        #             SUM(nominal) as nominal_total
        #         FROM trans_sales_order_paid_payment
        #         {where}
        #         GROUP BY company_id,cabang_id"""

        sql = f"""
                SELECT SUM
                ( amount_total ) - SUM ( amount_total_outstanding ) AS nominal_total
                FROM
                trans_inventory_subsidiary_invoice aa
                LEFT JOIN trans_inventory_subsidiary_sales_order_header bb ON bb.id_trans = aa.id_trans_sales_order
                {where}
                {gb}
                """

        try:
            print(sql)
            result = await self.db.executeToDict(sql)
            # print(result)
            if len(result) == 0:
                return 0
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return result[0]["nominal_total"]

    async def get_sales_month(
        self, company_id, cabang_id, tanggal, tanggal_start, is_pusat
    ):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year

        and_filter = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id}"""
        if int(company_id) == 1:
            and_filter = ""
        elif int(company_id) != 1 and is_pusat == True:
            and_filter = f"""WHERE company_id = {company_id}"""

        sql_chart_label = f""" SELECT
                            to_char(to_date(month_::text, 'MM'), 'Mon') AS month_name
                        FROM
                            (
                            SELECT
                                date_part( 'month', date_ ) AS month_,
                                date_part( 'year', date_ ) AS year_ 
                            FROM
                                ( SELECT generate_series ( '{tanggal_start}' :: DATE, '{tanggal}' :: DATE, '1 day' :: INTERVAL ) AS date_ ) AS dates 
                            ) A 
                        GROUP BY
                            month_,
                            year_ 
                        ORDER BY
                            year_, month_"""

        sql_chart_value = f"""SELECT 
                        B.total,
                        C.paid_paymnet
                    FROM
                        (
                        SELECT
                            month_,
                            year_ 
                        FROM
                            (
                            SELECT
                                date_part( 'month', date_ ) AS month_,
                                date_part( 'year', date_ ) AS year_ 
                            FROM
                                ( SELECT generate_series ( '{tanggal_start}' :: DATE, '{tanggal}' :: DATE, '1 day' :: INTERVAL ) AS date_ ) AS dates 
                            ) A 
                        GROUP BY
                            month_,
                            year_ 
                        ORDER BY
                            year_,
                            month_ 
                        )
                        A LEFT JOIN ( 
                        SELECT SUM (AA.amount) AS total,
                            date_part('month', aa.tanggal_invoice) as  MONTH,
                            date_part('year', aa.tanggal_invoice) as year FROM
                            trans_inventory_subsidiary_invoice aa
                            LEFT JOIN trans_inventory_subsidiary_sales_order_header bb ON bb.id_trans = aa.id_trans_sales_order
                            {and_filter}
                            GROUP BY month, year
                        ) B ON A.month_ = B.MONTH AND A.year_ = B.year
                        LEFT JOIN (
                            SELECT 
                            SUM( aa.amount_total) - sum(aa.amount_total_outstanding) as paid_paymnet,
                            date_part( 'month', payment_last_updated ) AS MONTH,
                            date_part( 'year', payment_last_updated ) AS year 
                            FROM
                            trans_inventory_subsidiary_invoice aa
                            LEFT JOIN trans_inventory_subsidiary_sales_order_header bb ON bb.id_trans = aa.id_trans_sales_order 
                            {and_filter}
                            GROUP BY
                            MONTH, year
                        ) C ON A.month_ = C.MONTH AND A.year_ = C.year
                        """
        print(sql_chart_value)
        try:
            result1 = await self.db.executeToDict(sql_chart_label)
            result_label = [kategori["month_name"] for kategori in result1]
            result2 = await self.db.executeToDict(sql_chart_value)
            result_value = [value["total"] for value in result2]
            result_paid = [value["paid_paymnet"] for value in result2]

            # print(sql_chart_label)

            data = {
                "result_label": result_label,
                "result_value": result_value,
                "result_paid": result_paid,
            }

            message = {"status": "success"}

        except Exception as e:
            message = {"status": "error"}
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def get_outstanding_payment_resume(
        self, tanggal, company_id, cabang_id, is_pusat
    ):
        # date =  datetime.strptime(tanggal, "%Y-%m-%d")

        and_filter = f"""and company_id = {company_id} AND cabang_id = {cabang_id}"""
        if int(company_id) == 1:
            and_filter = ""
        elif int(company_id) != 1 and is_pusat == True:
            and_filter = f"""and company_id = {company_id}"""

        sql = f"""SELECT
                    sum(amount_total_outstanding) total_out_standing,
                    sum(case when '{tanggal}'::DATE - aa.tanggal_due_date <= 0 then amount_total_outstanding else 0 end) no_due_date,
                    sum(case when '{tanggal}'::DATE - aa.tanggal_due_date BETWEEN 1 and 30 then amount_total_outstanding else 0 end) _1_30,
                    sum(case when '{tanggal}'::DATE - aa.tanggal_due_date BETWEEN 31 and 60 then amount_total_outstanding else 0 end) _31_60,
                    sum(case when '{tanggal}'::DATE - aa.tanggal_due_date BETWEEN 60 and 90 then amount_total_outstanding else 0 end) _60_90,
                    sum(case when '{tanggal}'::DATE - aa.tanggal_due_date > 90 then amount_total_outstanding else 0 end) _90
                FROM
                    trans_inventory_subsidiary_invoice aa
                    LEFT JOIN master_customer bb on aa.customer_id=bb.id_customer
                    LEFT JOIN master_company_cabang cc on bb.cabang_id=cc.id_cabang
                    LEFT JOIN master_company dd on bb.company_id=dd.id_company
                    WHERE complete_payment=FALSE {and_filter}"""

        try:
            print(sql)
            result = await self.db.executeToDict(sql)
            # print(result)

            data = {
                "total_out_standing": result[0]["total_out_standing"],
                "no_due_date": result[0]["no_due_date"],
                "_1_30": result[0]["_1_30"],
                "_31_60": result[0]["_31_60"],
                "_60_90": result[0]["_60_90"],
                "_90": result[0]["_90"],
            }

            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def read_outstanding_payment_detail(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        tanggal=None,
        is_pusat=False,
        filter_other="",
        filter_other_conj="",
    ):
        # print("tanggal ="+str(tanggal))

        where = f"""and company_id = {company_id} AND cabang_id = {cabang_id}"""
        if int(company_id) == 1:
            where = ""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""and company_id = {company_id}"""

        if tanggal != None and tanggal != "":
            data_tanggal = tanggal
            filter_other = f""
            filter_other_conj = f""
        else:
            data_tanggal = datetime.now().strftime("%Y-%m-%d")
            # print(data_tanggal)
            filter_other = f""
            filter_other_conj = f""
        if orderby == None or orderby == "":
            orderby = "zz.id_company,zz.cabang_name,zz.id_customer ASC"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""SELECT * FROM (SELECT
                    dd.id_company as id_company,
                    dd.company_name,
                    cc.id_cabang as id_cabang,
                    cc.cabang_name,
                    bb.id_customer as id_customer,
                    bb.nama_customer,
                    sum(amount_total_outstanding),
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date <= 0 then amount_total_outstanding else 0 end) no_due_date,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 1 and 30 then amount_total_outstanding else 0 end) _1_30,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 31 and 60 then amount_total_outstanding else 0 end) _31_60,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 60 and 90 then amount_total_outstanding else 0 end) _60_90,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date > 90 then amount_total_outstanding else 0 end) _90
                    FROM
                        trans_inventory_subsidiary_invoice aa
                        LEFT JOIN master_customer bb on aa.customer_id=bb.id_customer
                        LEFT JOIN master_company_cabang cc on bb.cabang_id=cc.id_cabang
                        LEFT JOIN master_company dd on bb.company_id=dd.id_company
                    WHERE complete_payment=FALSE {where}
                  	GROUP BY
                    dd.id_company,
                    dd.company_name,
                    cc.id_cabang,
                    cc.cabang_name,
                    bb.id_customer,
                    bb.nama_customer) zz"""
            + str_clause
        )

        sql_count = (
            f"""SELECT count(zz.*) count FROM 
                    (SELECT * FROM (SELECT
                    dd.id_company as id_company,
                    dd.company_name,
                    cc.id_cabang as id_cabang,
                    cc.cabang_name,
                    bb.id_customer as id_customer,
                    bb.nama_customer,
                    sum(amount_total_outstanding),
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date <= 0 then amount_total_outstanding else 0 end) no_due_date,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 1 and 60 then amount_total_outstanding else 0 end) _1_60,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 61 and 180 then amount_total_outstanding else 0 end) _61_120,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 181 and 365 then amount_total_outstanding else 0 end) _181_365,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date > 365 then amount_total_outstanding else 0 end) _365
                    FROM
                        trans_inventory_subsidiary_invoice aa
                        LEFT JOIN master_customer bb on aa.customer_id=bb.id_customer
                        LEFT JOIN master_company_cabang cc on bb.cabang_id=cc.id_cabang
                        LEFT JOIN master_company dd on bb.company_id=dd.id_company
                    WHERE complete_payment=FALSE {where}
                  	GROUP BY
                    dd.id_company,
                    dd.company_name,
                    cc.id_cabang,
                    cc.cabang_name,
                    bb.id_customer,
                    bb.nama_customer) zz ) zz"""
            + str_clause_count
        )

        # print(sql)

        try:
            result = await self.db.executeToDict(sql)
            print(result)
            result_count = await self.db.executeToDict(sql_count)

            data = {"data": result, "count": result_count[0]["count"]}

        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def export_outstanding_payment(
        self,
        company_id,
        cabang_id,
        data_tanggal,
        is_pusat,
    ):
        where = f"""and company_id = {company_id} AND cabang_id = {cabang_id}"""
        if int(company_id) == 1:
            where = ""

        if int(company_id) != 1 and is_pusat == True:
            where = f"""and company_id = {company_id}"""

        print(where)

        sql = f"""SELECT * FROM (SELECT
                    dd.company_name,
                    cc.id_cabang as id_cabang,
                    cc.cabang_name,
                    bb.nama_customer,
                    sum(amount_total_outstanding) as total_outstanding,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date <= 0 then amount_total_outstanding else 0 end) no_due_date,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 1 and 21 then amount_total_outstanding else 0 end) _1_21,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 22 and 60 then amount_total_outstanding else 0 end) _22_60,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date BETWEEN 60 and 90 then amount_total_outstanding else 0 end) _60_90,
                    sum(case when '{data_tanggal}'::DATE - aa.tanggal_due_date > 90 then amount_total_outstanding else 0 end) _90
                    FROM
                        trans_inventory_subsidiary_invoice aa
                        LEFT JOIN master_customer bb on aa.customer_id=bb.id_customer
                        LEFT JOIN master_company_cabang cc on bb.cabang_id=cc.id_cabang
                        LEFT JOIN master_company dd on bb.company_id=dd.id_company
                    WHERE complete_payment=FALSE {where}
                  	GROUP BY
                    dd.id_company,
                    dd.company_name,
                    cc.id_cabang,
                    cc.cabang_name,
                    bb.id_customer,
                    bb.nama_customer) zz"""

        try:
            result = await self.db.executeToDict(sql)
            # print(sql)
            wb = self.generate_excel_outstanding_payment(result)
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=example.xlsx"},
            )

        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

    def generate_excel_outstanding_payment(self, result_data):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "Company Name"
        ws["B1"].value = "Cabang ID"
        ws["C1"].value = "Cabang Name"
        ws["D1"].value = "Nama Customer"
        ws["E1"].value = "Total Outstanding"
        ws["F1"].value = "No Due Date"
        ws["G1"].value = "1 - 21"
        ws["H1"].value = "22 - 60"
        ws["I1"].value = "61 - 90"
        ws["J1"].value = "> 90"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb

    async def read_inventori_resume(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        is_pusat,
        filter_other="",
        filter_other_conj="",
    ):
        # print("tanggal ="+str(tanggal))
        if orderby == None or orderby == "":
            orderby = "zz.produk_id ASC"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        where = f"""Where company_id = {company_id} AND cabang_id = {cabang_id}"""
        if int(company_id) == 1:
            where = ""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""where company_id = {company_id}"""

        query = f"""SELECT * FROM
                (SELECT
                    xx.produk_id,
                    xx.nama_produk,	
                    xx.qty,
                    CASE 
                        WHEN xx.qty <> 0 
                    THEN Round(xx.harga_total/xx.qty,2)::FLOAT
                    ELSE 0
                    END AS harga_sat_hpp,
                    xx.harga_total
                FROM 
                (
                SELECT 
                    aa.produk_id,
                    dd.nama_produk,
                    SUM(aa.qty) as qty,
                    SUM(aa.harga_satuan) as harga_satuan,
                    SUM(aa.harga_total) as harga_total
                FROM trans_inventory_detail aa
                LEFT JOIN master_produk dd on aa.produk_id = dd.id_produk
                {where}
                GROUP BY aa.produk_id,dd.nama_produk
                ) xx
            )zz"""

        sql = query + str_clause

        sql_2 = query + str_clause_count
        sql_count = (
            sql_count
        ) = f"""SELECT COUNT(*) 
        FROM ({sql_2})  as subquery"""

        try:
            result = await self.db.executeToDict(sql)
            result_count = await self.db.executeToDict(sql_count)

            data = {"data": result, "count": result_count[0]["count"]}

        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def get_biaya_admin(
        self, tanggal, company_id, cabang_id, tanggal_start, is_pusat
    ):
        date = datetime.strptime(tanggal, "%Y-%m-%d")
        tahun = date.year
        where = f"""WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        if int(company_id) == 1:
            where = f"""WHERE tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE company_id = {company_id} AND tanggal BETWEEN '{tanggal_start}' AND '{tanggal}'"""

        sql = f"""
                    SELECT sum(b.biaya_admin) as total from trans_inventory_subsidiary_invoice a 
                    LEFT JOIN trans_inventory_subsidiary_sales_order_header b on a.id_trans_sales_order = b.id_trans
                    {where}
        """

        try:
            result = await self.db.executeToDict(sql)

        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))

        return result

    async def read_data_all(
        self, company_id, cabang_id, tanggal, tanggal_start, is_pusat
    ):

        data = {
            "inventori": await self.get_inventori(
                company_id, cabang_id, tanggal, tanggal_start, is_pusat
            ),
            "sales": await self.get_sales(
                company_id, cabang_id, tanggal, tanggal_start, is_pusat
            ),
            "sales_month": await self.get_sales_month(
                company_id, cabang_id, tanggal, tanggal_start, is_pusat
            ),
            "outstanding_payment": await self.get_outstanding_payment(
                company_id, cabang_id, tanggal, tanggal_start, is_pusat
            ),
            "paid_payment": await self.get_paid_payment(
                company_id, cabang_id, tanggal, tanggal_start, is_pusat
            ),
            "outstanding_payment_resume": await self.get_outstanding_payment_resume(
                tanggal, company_id, cabang_id, is_pusat
            ),
            "biaya_admin": await self.get_biaya_admin(
                tanggal, company_id, cabang_id, tanggal_start, is_pusat
            ),
        }
        print("\n\n\n", data)

        return data

    async def read_sales_header(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        tanggal=None,
        tanggal_start=None,
        is_pusat=False,
        filter_other="",
        filter_other_conj="",
    ):

        where = f"""WHERE b.company_id = {company_id} AND b.cabang_id = {cabang_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}' and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null) """
        if int(company_id) == 1:
            where = f"WHERE tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null)"
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE b.company_id = {company_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null)"""

        if orderby == None or orderby == "":
            orderby = "company_id, cabang_id, produk_id"

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        query = f"""
                SELECT C.company_name,
        D.cabang_name,
        B.nama_produk,
        A.total_qty,
        A.total_sales,
        A.total_hpp,
        A.total_sales - A.total_hpp as margin,
        ROUND((A.total_sales - A.total_hpp) / A.total_sales * 100, 2) as margin_percent,
        A.order_type
        FROM
        (
            SELECT
            B.company_id,
            B.cabang_id,
            C.produk_id,
            SUM ( C.qty ) AS total_qty,
            SUM ( C.harga_total ) AS total_sales,
            SUM(c.harga_total_hpp) as total_hpp,
            B.order_type
            FROM
            trans_inventory_subsidiary_invoice A
            LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans     
            LEFT JOIN trans_inventory_subsidiary_sales_order C ON B.id_trans = C.id_trans
            {where}
            GROUP BY
            B.order_type,
            c.produk_id,
            B.company_id,
            B.cabang_id
        )
        A LEFT JOIN master_produk B ON A.produk_id = B.id_produk
        LEFT JOIN master_company C ON A.company_id = C.id_company
        LEFT JOIN master_company_cabang D ON A.company_id = D.id_company
        AND A.cabang_id = D.id_cabang
        """
        sql = query + str_clause
        # print(sql)

        sql_2 = query + str_clause_count

        sql_count = f"""SELECT COUNT(*) 
        FROM ({sql_2})  as subquery"""
        try:
            result = await self.db.executeToDict(sql)
            result_count = await self.db.executeToDict(sql_count)
            data = {"data": result, "total": result_count[0]["count"]}

        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def export_sales_header(
        self,
        company_id,
        cabang_id,
        tanggal=None,
        tanggal_start=None,
        is_pusat=False,
    ):
        where = f"""WHERE b.company_id = {company_id} AND b.cabang_id = {cabang_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}' and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null) """
        if int(company_id) == 1:
            where = f"WHERE tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null)"
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE b.company_id = {company_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and b.status_release = true AND (b.approval_status !=4 or b.approval_status is null)"""

        orderby = "company_id, cabang_id, produk_id"

        sql = f"""
                SELECT C.company_name,
        D.cabang_name,
        B.nama_produk,
        A.order_type,
        A.total_qty,
        A.total_sales,
        A.total_hpp,
        A.total_sales - A.total_hpp as margin,
        ROUND((A.total_sales - A.total_hpp) / A.total_sales * 100, 2) as margin_percent
        FROM
        (
            SELECT
            B.company_id,
            B.cabang_id,
            C.produk_id,
            B.order_type,
            SUM ( C.qty ) AS total_qty,
            SUM ( C.harga_total ) AS total_sales,
            SUM(c.harga_total_hpp) as total_hpp
            FROM
            trans_inventory_subsidiary_invoice A
            LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans     
            LEFT JOIN trans_inventory_subsidiary_sales_order C ON B.id_trans = C.id_trans
            {where}
            GROUP BY
            B.order_type,
            c.produk_id,
            B.company_id,
            B.cabang_id
        )
        A LEFT JOIN master_produk B ON A.produk_id = B.id_produk
        LEFT JOIN master_company C ON A.company_id = C.id_company
        LEFT JOIN master_company_cabang D ON A.company_id = D.id_company
        AND A.cabang_id = D.id_cabang
        """

        # print(sql)

        try:
            result = await self.db.executeToDict(sql)
            # print(sql)
            wb = self.generate_excel_sales_header(result)
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=example.xlsx"},
            )
        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

        return result

    def generate_excel_sales_header(self, result_data):
        wb = Workbook()
        ws = wb.active

        ws["A1"].value = "Company Name"
        ws["B1"].value = "Cabang Name"
        ws["C1"].value = "Nama Produk"
        ws["D1"].value = "Order Type"
        ws["E1"].value = "Total Qty"
        ws["F1"].value = "Total Sales"
        ws["G1"].value = "HPP"
        ws["H1"].value = "Margin"
        ws["I1"].value = "Margin Percent"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb

    async def read_sales_header_produk(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        tanggal=None,
        tanggal_start=None,
        is_pusat=False,
        filter_other="",
        filter_other_conj="",
    ):
        where = f"""WHERE d.company_id = {company_id} AND d.cabang_id = {cabang_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}' and d.status_release = true AND (approval_status !=4 or approval_status is null)"""
        if int(company_id) == 1:
            where = f"WHERE tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and d.status_release = true AND (approval_status !=4 or approval_status is null)"
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE d.company_id = {company_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and d.status_release = true AND (approval_status !=4 or approval_status is null)"""

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        query = f"""
            SELECT B.nama_produk, A.*, ROUND(A.margin / A.amount_total * 100,2) as margin_percent FROM (
                SELECT 
                        A.produk_id,
                        d.order_type,
                        SUM ( A.qty ) total_qty,
                        SUM ( A.harga_total ) AS amount_total,
                        SUM ( A.harga_total_hpp ) AS hpp,
                        SUM ( A.harga_total ) - SUM ( A.harga_total_hpp ) AS margin
                    FROM trans_inventory_subsidiary_invoice Z
                    LEFT JOIN trans_inventory_subsidiary_sales_order_header D on Z.id_trans_sales_order = D.id_trans
                    LEFT JOIN trans_inventory_subsidiary_sales_order A on D.id_trans = A.id_trans
                    LEFT JOIN master_company B on A.company_id = B.id_company
                    LEFT JOIN master_company_cabang C on A.cabang_id = C.id_cabang and A.company_id = C.id_company
                    {where}
                    GROUP BY D.order_type, A.produk_id
                ) A
                LEFT JOIN master_produk B on A.produk_id = B.id_produk
        """

        sql = query + str_clause

        sql_2 = query + str_clause_count

        sql_count = f"""SELECT COUNT(*) 
        FROM ({sql_2})  as subquery"""

        # print(sql)

        try:
            result = await self.db.executeToDict(sql)
            result_count = await self.db.executeToDict(sql_count)
            data = {"data": result, "total": result_count[0]["count"]}

        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

        return data

    async def export_sales_header_produk(
        self,
        company_id,
        cabang_id,
        tanggal=None,
        tanggal_start=None,
        is_pusat=False,
    ):
        where = f"""WHERE d.company_id = {company_id} AND d.cabang_id = {cabang_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}' and d.status_release = true AND (approval_status !=4 or approval_status is null)"""
        if int(company_id) == 1:
            where = f"WHERE tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and d.status_release = true AND (approval_status !=4 or approval_status is null)"
        elif int(company_id) != 1 and is_pusat == True:
            where = f"""WHERE d.company_id = {company_id} and tanggal_invoice between '{tanggal_start}' and '{tanggal}'  and d.status_release = true AND (approval_status !=4 or approval_status is null)"""

        sql = f"""
            SELECT B.nama_produk, A.*, ROUND(A.margin / A.amount_total * 100,2) as margin_percent FROM (
                SELECT 
                    A.produk_id,
                    d.order_type,
                    SUM ( A.qty ) total_qty,
                    SUM ( A.harga_total ) AS amount_total,
                    SUM ( A.harga_total_hpp ) AS hpp,
                    SUM ( A.harga_total ) - SUM ( A.harga_total_hpp ) AS margin
                FROM trans_inventory_subsidiary_invoice Z
                LEFT JOIN trans_inventory_subsidiary_sales_order_header D on Z.id_trans_sales_order = D.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order A on D.id_trans = A.id_trans
                LEFT JOIN master_company B on A.company_id = B.id_company
                LEFT JOIN master_company_cabang C on A.cabang_id = C.id_cabang and A.company_id = C.id_company
                {where}
                GROUP BY D.order_type, A.produk_id
            ) A
            LEFT JOIN master_produk B on A.produk_id = B.id_produk
        """

        try:
            result = await self.db.executeToDict(sql)
            # print(sql)
            wb = self.generate_excel_sales_header_produk(result)
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=example.xlsx"},
            )
        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

        return result

    def generate_excel_sales_header_produk(self, result_data):
        wb = Workbook()
        ws = wb.active

        ws["A1"].value = "Nama Produk"
        ws["B1"].value = "Produk Id"
        ws["C1"].value = "Order Type"
        ws["D1"].value = "Total Qty"
        ws["E1"].value = "Amount Total"
        ws["F1"].value = "HPP"
        ws["G1"].value = "Margin"
        ws["H1"].value = "Margin Percent"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb


"""
list your path url at bottom
example /testing url
test from postman :
url/api/c_dashboard_utama/testing
for post method and other method, check tutorial from 
https://fastapi.tiangolo.com/
"""


@app.get("/api/f_dashboard/c_dashboard_utama/read_outstanding_payment_detail")
async def read_outstanding_payment_detail(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    tanggal: str = Query(None, alias="$tanggal"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.read_outstanding_payment_detail(
        orderby, limit, offset, filter, company_id, cabang_id, tanggal, is_pusat
    )


@app.get("/api/f_dashboard/c_dashboard_utama/export_outstanding_payment")
async def read_outstanding_payment_detail(
    tanggal: str = Query(None, alias="tanggal"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.export_outstanding_payment(
        company_id, cabang_id, tanggal, is_pusat
    )


@app.get("/api/f_dashboard/c_dashboard_utama/read_inventori_resume")
async def read_inventori_resume(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.read_inventori_resume(
        orderby, limit, offset, filter, company_id, cabang_id, is_pusat
    )


@app.get("/api/f_dashboard/c_dashboard_utama/read_data_all")
async def read_data_all(
    company_id, cabang_id, tanggal: str, tanggal_start: str, is_pusat: bool
):
    ob_data = c_dashboard_utama()
    return await ob_data.read_data_all(
        company_id, cabang_id, tanggal, tanggal_start, is_pusat
    )


@app.get("/api/f_dashboard/c_dashboard_utama/read_sales_header")
async def read_sales_header(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    tanggal: str = Query(None, alias="tanggal"),
    tanggal_start: str = Query(None, alias="tanggal_start"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.read_sales_header(
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        tanggal,
        tanggal_start,
        is_pusat,
    )


@app.get("/api/f_dashboard/c_dashboard_utama/read_sales_header_produk")
async def read_sales_header_produk(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    tanggal: str = Query(None, alias="tanggal"),
    tanggal_start: str = Query(None, alias="tanggal_start"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.read_sales_header_produk(
        orderby,
        limit,
        offset,
        filter,
        company_id,
        cabang_id,
        tanggal,
        tanggal_start,
        is_pusat,
    )


@app.get("/api/f_dashboard/c_dashboard_utama/export_sales_header")
async def read_sales_header(
    tanggal: str = Query(None, alias="tanggal"),
    tanggal_start: str = Query(None, alias="tanggal_start"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.export_sales_header(
        company_id, cabang_id, tanggal, tanggal_start, is_pusat
    )


@app.get("/api/f_dashboard/c_dashboard_utama/export_sales_header_produk")
async def read_sales_header(
    tanggal: str = Query(None, alias="tanggal"),
    tanggal_start: str = Query(None, alias="tanggal_start"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_pusat: bool = Query(None, alias="is_pusat"),
):
    ob_data = c_dashboard_utama()
    return await ob_data.export_sales_header_produk(
        company_id, cabang_id, tanggal, tanggal_start, is_pusat
    )
