import base64
from datetime import date, datetime
import io

from fastapi import Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


from library import *
import os
from library.router import app
from library.db import Db
from pydantic import BaseModel


class c_salesman_report(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self,
        orderby,
        limit,
        offset,
        filter,
        salesman=None,
        company_id=None,
        cabang_id=None,
        tanggal=None,
        filter_other="",
        filter_other_conj="",
    ):
        month = datetime.strptime(tanggal, "%Y-%m-%d").month
        year = datetime.strptime(tanggal, "%Y-%m-%d").year

        salesman_filter = f"and salesman = '{salesman}'"

        cabang_filter = f" AND cabang_id = '{cabang_id}'"
        if cabang_id == 0:
            cabang_filter = ""

        if salesman == 0:
            salesman_filter = f""

        filter_other = f" company_id = '{company_id}' {cabang_filter} {salesman_filter}"
        filter_other_conj = f" and "

        # print(filter_other)
        if orderby == None or orderby == "":
            orderby = "company_id, cabang_id, name"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        query = (
            f"""
            SELECT
            B.username,
                B.NAME,
                B.company_id,
                B.cabang_id,
                C.company_name,
                D.cabang_name,
                A.*
                FROM
                (
                    SELECT
                    b.salesman,
                    SUM ( A.amount_total ) AS total_sales,
                    SUM ( CASE WHEN A.tanggal_invoice BETWEEN '{year}-{month}-01' AND '{tanggal}' THEN A.amount_total ELSE 0 END ) AS monthly_sales,
                    SUM ( A.amount_total_outstanding ) AS total_outstanding,
                SUM ( CASE WHEN '{tanggal}' :: DATE - A.tanggal_due_date <= 0 THEN amount_total_outstanding ELSE 0 END ) no_due_date,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date between 1 and 15  THEN amount_total_outstanding ELSE 0 END ) AS overdue_less_15,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 15 THEN amount_total_outstanding ELSE 0 END ) AS overdue_more_15,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 0 THEN amount_total_outstanding ELSE 0 END ) AS total_overdue,
                SUM ( A.amount_total - A.amount_total_outstanding ) AS total_paid 
                FROM
                trans_inventory_subsidiary_invoice
                A LEFT JOIN trans_inventory_subsidiary_sales_order b ON A.id_trans_sales_order = b.id_trans 
                
                WHERE
                b.status_release = TRUE 
                GROUP BY
                b.salesman 
                )
                A LEFT JOIN master_user B ON A.salesman = B.id_user
                LEFT JOIN master_company c ON c.id_company = b.company_id 
                LEFT JOIN master_company_cabang d ON d.id_cabang = b.cabang_id and d.id_company = b.company_id
        """
            + str_clause
        )

        query_count = (
            f"""
           SELECT
            COUNT(*)
            FROM
            (
                SELECT
                b.salesman,
                SUM ( A.amount_total ) AS total_sales,
                SUM ( CASE WHEN A.tanggal_invoice BETWEEN '2025-10-01' AND '2025-10-30' THEN A.amount_total ELSE 0 END ) AS month_sales,
                SUM ( A.amount_total_outstanding ) AS total_outstanding,
            SUM ( CASE WHEN '{tanggal}' :: DATE - A.tanggal_due_date <= 0 THEN amount_total_outstanding ELSE 0 END ) no_due_date,
            SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date between 1 and 15  THEN amount_total_outstanding ELSE 0 END ) AS overdue_less_15,
            SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 15 THEN amount_total_outstanding ELSE 0 END ) AS overdue_more_15,
            SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 0 THEN amount_total_outstanding ELSE 0 END ) AS total_overdue,
            SUM ( A.amount_total - A.amount_total_outstanding ) AS total_paid 
            FROM
            trans_inventory_subsidiary_invoice
            A LEFT JOIN trans_inventory_subsidiary_sales_order b ON A.id_trans_sales_order = b.id_trans 
            WHERE
            b.status_release = TRUE 
            GROUP BY
            b.salesman 
            )
            A LEFT JOIN master_user B ON A.salesman = B.id_user
          """
            + str_clause_count
        )

        print(query)
        # print(query_count)

        result = await self.db.executeToDict(query)
        result_count = await self.db.executeToDict(query_count)
        data = {"data": result, "total": result_count[0]["count"]}
        return data

    async def invoice_per_salesman(
        self,
        orderby,
        limit,
        offset,
        filter,
        salesman=None,
        company_id=None,
        cabang_id=None,
        tanggal=None,
        index=None,
        filter_other="",
        filter_other_conj="",
    ):

        index_filter = f" '{tanggal}':: DATE - A.tanggal_due_date <= 0"

        if index == 2:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date BETWEEN 1 and 15"
        elif index == 3:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date > 15"
        elif index == 4:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date > 0"

        where = f"""
        {index_filter}
        AND b.cabang_id = {cabang_id} 
        AND b.company_id = {company_id}
        AND b.salesman = {salesman} 
        AND b.status_release = TRUE
        AND a.complete_payment = False
        """

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )

        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""
            SELECT A
        .id_trans,
        D.nama_customer,
        A.amount_total,
        A.amount_total_outstanding,
        
        A.tanggal_due_date,
        D.no_ktp,
        D.alamat,
        D.npwp 
        FROM
        trans_inventory_subsidiary_invoice
        A LEFT JOIN trans_inventory_subsidiary_sales_order B ON A.id_trans_sales_order = b.id_trans
        LEFT JOIN master_produk C ON A.produk_id = C.id_produk
        LEFT JOIN master_customer D ON B.customer_id = D.id_customer 
        WHERE
        {where}
        """
            + str_clause
        )

        sql_count = (
            f"""
        SELECT
        COUNT(*)
        FROM
        trans_inventory_subsidiary_invoice
        A LEFT JOIN trans_inventory_subsidiary_sales_order B ON A.id_trans_sales_order = b.id_trans
        LEFT JOIN master_produk C ON A.produk_id = C.id_produk
        LEFT JOIN master_customer D ON B.customer_id = D.id_customer 
        WHERE
        {where}
        """
            + str_clause_count
        )

        print(sql)

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)
        data = {"data": result, "total": result_count[0]["count"]}
        return data

    async def export_excel(
        self,
        salesman=None,
        company_id=None,
        cabang_id=None,
        tanggal=None,
    ):

        month = datetime.strptime(tanggal, "%Y-%m-%d").month
        year = datetime.strptime(tanggal, "%Y-%m-%d").year

        salesman_filter = f"and salesman = '{salesman}'"

        cabang_filter = f" AND cabang_id = '{cabang_id}'"
        if cabang_id == 0:
            cabang_filter = ""

        if salesman == 0:
            salesman_filter = f""

        where = f"where company_id = '{company_id}' {cabang_filter} {salesman_filter}"

        sql = f"""
            SELECT
            B.username,
                B.NAME,
                B.company_id,
                B.cabang_id,
                C.company_name,
                D.cabang_name,
                A.*
                FROM
                (
                    SELECT
                    b.salesman,
                    SUM ( A.amount_total ) AS total_sales,
                    SUM ( CASE WHEN A.tanggal_invoice BETWEEN '{year}-{month}-01' AND '{tanggal}' THEN A.amount_total ELSE 0 END ) AS monthly_sales,
                    SUM ( A.amount_total_outstanding ) AS total_outstanding,
                SUM ( CASE WHEN '{tanggal}' :: DATE - A.tanggal_due_date <= 0 THEN amount_total_outstanding ELSE 0 END ) no_due_date,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date between 1 and 15  THEN amount_total_outstanding ELSE 0 END ) AS overdue_less_15,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 15 THEN amount_total_outstanding ELSE 0 END ) AS overdue_more_15,
                SUM ( CASE WHEN '{tanggal}':: DATE - A.tanggal_due_date > 0 THEN amount_total_outstanding ELSE 0 END ) AS total_overdue,
                SUM ( A.amount_total - A.amount_total_outstanding ) AS total_paid 
                FROM
                trans_inventory_subsidiary_invoice
                A LEFT JOIN trans_inventory_subsidiary_sales_order b ON A.id_trans_sales_order = b.id_trans 
                
                WHERE
                b.status_release = TRUE 
                GROUP BY
                b.salesman 
                )
                A LEFT JOIN master_user B ON A.salesman = B.id_user
                LEFT JOIN master_company c ON c.id_company = b.company_id 
                LEFT JOIN master_company_cabang d ON d.id_cabang = b.cabang_id and d.id_company = b.company_id
                {where}
                ORDER BY B.company_id, B.cabang_id, A.salesman
        """
        result = await self.db.executeToDict(sql)
        print(sql)

        wb = self.excel_return_header(result, tanggal, month, year)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def excel_return_header(self, result_data, tanggal, month, year):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "Username"
        ws["B1"].value = "Nama Sales"
        ws["C1"].value = "Nama Company"
        ws["D1"].value = "Nama Cabang"
        ws["E1"].value = f"Total Sales S/d{tanggal}"
        ws["F1"].value = f"Sales Bulan {month}-{year}"
        ws["G1"].value = "Total Outstanding"
        ws["H1"].value = "No Due Date"
        ws["I1"].value = "Overdue < 15"
        ws["J1"].value = "Overdue > 15"
        ws["K1"].value = "Total Overdue"
        ws["L1"].value = "Total Paid"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for index, (key, value) in enumerate(result_data[0].items()):
            if index == 2 or index == 3 or index == 6:
                continue
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        print(data_key)

        for data in result_data:
            data_export = []
            for index, key in enumerate(data_key):
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb

    async def export_excel_invoice(
        self,
        salesman=None,
        company_id=None,
        cabang_id=None,
        tanggal=None,
        index=None,
    ):
        index_filter = f" '{tanggal}':: DATE - A.tanggal_due_date <= 0"
        print(index)

        if index == 2:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date BETWEEN 1 and 15"
        elif index == 3:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date > 15"
        elif index == 4:
            index_filter = f" '{tanggal}' :: DATE - A.tanggal_due_date > 0"

        where = f"""
        {index_filter}
        AND b.cabang_id = {cabang_id}
        AND b.company_id = {company_id}
        AND b.salesman = {salesman}
        AND b.status_release = TRUE
        """

        query = f"""
            SELECT A
        .id_trans,
        D.nama_customer,
        A.amount_total,
        A.amount_total_outstanding,

        A.tanggal_due_date,
        D.no_ktp,
        D.alamat,
        D.npwp
        FROM
        trans_inventory_subsidiary_invoice
        A LEFT JOIN trans_inventory_subsidiary_sales_order B ON A.id_trans_sales_order = b.id_trans
        LEFT JOIN master_produk C ON A.produk_id = C.id_produk
        LEFT JOIN master_customer D ON B.customer_id = D.id_customer
        WHERE
        {where}
        """

        print(query)

        result = await self.db.executeToDict(query)

        wb = self.excel_return_invoice(result)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def excel_return_invoice(self, result_data):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "No Invoice"
        ws["B1"].value = "Nama Customer"
        ws["C1"].value = "Total"
        ws["D1"].value = "Total Outstanding"
        ws["E1"].value = "Tanggal Due Date"
        ws["F1"].value = "No KTP"
        ws["G1"].value = "Alamat"
        ws["H1"].value = "NPWP"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            print(key, value)
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb


@app.get("/api/f_report/c_salesman_report/read")
async def test_get(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    salesman: int = Query(None, alias="salesman"),
    tanggal: str = Query(None, alias="tanggal"),
):
    ob_data = c_salesman_report()
    return await ob_data.read(
        orderby, limit, offset, filter, salesman, company_id, cabang_id, tanggal
    )


@app.get("/api/f_report/c_salesman_report/read_invoice")
async def test_get(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    salesman: int = Query(None, alias="salesman"),
    tanggal: str = Query(None, alias="tanggal"),
    index: int = Query(None, alias="index"),
):
    ob_data = c_salesman_report()
    return await ob_data.invoice_per_salesman(
        orderby, limit, offset, filter, salesman, company_id, cabang_id, tanggal, index
    )


@app.get("/api/f_report/c_salesman_report/export_excel")
async def test_get(
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    salesman: int = Query(None, alias="salesman"),
    tanggal: str = Query(None, alias="tanggal"),
):
    ob_data = c_salesman_report()
    return await ob_data.export_excel(salesman, company_id, cabang_id, tanggal)


@app.get("/api/f_report/c_salesman_report/export_invoice")
async def test_get(
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    salesman: int = Query(None, alias="salesman"),
    tanggal: str = Query(None, alias="tanggal"),
    index: int = Query(None, alias="index"),
):
    ob_data = c_salesman_report()
    return await ob_data.export_excel_invoice(
        salesman, company_id, cabang_id, tanggal, index
    )
