import base64
from datetime import datetime
import io

from fastapi import Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from library import *
import os
from library.router import app
from library.db import Db
from pydantic import BaseModel
from openpyxl.styles import PatternFill, Font


class c_margin_report(object):
    def __init__(self):
        self.db = Db()

    async def export_margin_report(
        self, tanggal_awal, tanggal_akhir, company_id, is_range
    ):
        company_id = str(company_id)
        month = datetime.strptime(tanggal_akhir, "%Y-%m-%d").month
        year = datetime.strptime(tanggal_akhir, "%Y-%m-%d").year

        where = f"Where complete_payment = true AND b.company_id = {company_id} AND "

        is_range_where = f"date_part('month', payment_last_updated) = {month} and date_part('year', payment_last_updated) = {year}"

        if is_range:
            is_range_where = f""" payment_last_updated BETWEEN '{tanggal_awal}' AND '{tanggal_akhir}' """

        where += is_range_where

        sql_margin_by_order_type = self.margin_by_order_type(where)
        sql_margin_by_produk = self.margin_by_produk(where)
        sql_margin_by_cabang = self.margin_by_cabang(where)
        sql_margin_by_invoice = self.margin_by_invoice(where)

        print(sql_margin_by_order_type)

        by_order_type = await self.db.executeToDict(sql_margin_by_order_type)
        by_produk = await self.db.executeToDict(sql_margin_by_produk)
        by_cabang = await self.db.executeToDict(sql_margin_by_cabang)
        by_invoice = await self.db.executeToDict(sql_margin_by_invoice)

        if is_range:
            sql_margin_by_order_type_per_month = self.margin_by_order_type_per_month(
                where
            )
            by_order_type_month = await self.db.executeToDict(
                sql_margin_by_order_type_per_month
            )
            sql_margin_by_cabang_month = self.margin_by_cabang_month(where)
            by_month = await self.db.executeToDict(sql_margin_by_cabang_month)

        wb = Workbook()
        sheets = {}
        header1 = [
            "Nama Company",
            "Order type",
            "Sales",
            "HPP",
            "Margin",
            "Margin Percent",
        ]
        sheets["BY ORDER TYPE"] = {
            "data": by_order_type,
            "header": header1,
        }

        if is_range:
            header6 = [
                "Nama Company",
                "Order type",
                "Month",
                "Year",
                "Sales",
                "HPP",
                "Margin",
                "Margin Percent",
            ]
            sheets["BY ORDER TYPE MONTHLY"] = {
                "data": by_order_type_month,
                "header": header6,
            }

        header2 = [
            "Nama Company",
            "Nama Produk",
            "Qty",
            "Order type",
            "Sales",
            "HPP",
            "Margin",
            "Margin Percent",
        ]
        sheets["BY PRODUK "] = {
            "data": by_produk,
            "header": header2,
        }

        header3 = [
            "Nama Company",
            "Nama Cabang",
            "Order type",
            "Sales",
            "HPP",
            "Margin",
            "Margin Percent",
        ]
        sheets["BY CABANG"] = {
            "data": by_cabang,
            "header": header3,
        }

        if is_range:
            header4 = [
                "Nama Company",
                "Nama Cabang",
                "Year",
                "Month",
                "Order type",
                "Sales",
                "HPP",
                "Margin",
                "Margin Percent",
            ]
            sheets["BY MONTH"] = {
                "data": by_month,
                "header": header4,
            }

        header5 = [
            "ID INVOICE",
            "Nama Company",
            "Nama Cabang",
            "Order Type",
            "Nama Produk",
            "Qty",
            "Qty in Kg",
            "HPP",
            "Sales",
            "Margin",
            "Margin Percent",
            "Tanggal Pelunasan",
        ]

        sheets["BY INVOICE"] = {
            "data": by_invoice,
            "header": header5,
        }

        wb.remove(wb.active)

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(sheet_name)
            self.write_dict_data(ws, data["data"], data["header"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def write_dict_data(self, ws, data, header):
        HEADER_FONT = Font(b=True, color="000000")
        HEADER_FILL = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        if not data:
            return

        keys = list(data[0].keys())

        # Header
        for col, key in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Rows
        for row in data:
            ws.append([row.get(key) for key in keys])

    def margin_by_order_type(self, where):
        sql = f"""
            SELECT  C.company_name, A.order_type, A.sales, A.hpp, A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent FROM (
  SELECT  SUM(B.harga_total) as sales, SUM(c.hpp) as hpp, b.company_id, b.order_type FROM trans_inventory_subsidiary_invoice A
  LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
  LEFT JOIN (
    SELECT id_trans, SUM (harga_total) sales, SUM(harga_total_hpp) as hpp FROM trans_inventory_subsidiary_sales_order A
    GROUP BY id_trans 
  ) C ON B.id_trans = C.id_trans
  {where}
  GROUP BY B.company_id, B.order_type
) A
LEFT JOIN master_company C on A.company_id = C.id_company
ORDER BY company_id;
                """

        return sql

    def margin_by_order_type_per_month(self, where):
        sql = f"""
        SELECT  C.company_name, A.order_type, A.month, A.year, A.sales, A.hpp, A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent FROM (
  SELECT  SUM(B.harga_total) as sales, SUM(c.hpp) as hpp, b.company_id, b.order_type, date_part('year', payment_last_updated) as year, date_part('month', payment_last_updated) as month FROM trans_inventory_subsidiary_invoice A
  LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
  LEFT JOIN (
    SELECT id_trans, SUM (harga_total) sales, SUM(harga_total_hpp) as hpp FROM trans_inventory_subsidiary_sales_order A
    GROUP BY id_trans 
  ) C ON B.id_trans = C.id_trans
  {where}
  GROUP BY date_part('year', payment_last_updated), date_part('month', payment_last_updated), B.company_id, B.order_type
) A
LEFT JOIN master_company C on A.company_id = C.id_company
ORDER BY company_id, month, year, order_type;
        """

        return sql

    def margin_by_produk(self, where):
        sql = f"""
            SELECT  C.company_name, B.nama_produk, A.qty, A.order_type, A.sales, A.hpp, A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent FROM (
  SELECT SUM(c.qty) as qty, SUM(c.harga_total) as sales, SUM(c.harga_total_hpp) as hpp, b.company_id, c.produk_id, b.order_type FROM trans_inventory_subsidiary_invoice A
  LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
  LEFT JOIN trans_inventory_subsidiary_sales_order C ON B.id_trans = C.id_trans
  {where}
  GROUP BY B.company_id, B.order_type, C.produk_id
) A
LEFT JOIN master_produk B on A.produk_id = B.id_produk
LEFT JOIN master_company C on A.company_id = C.id_company
ORDER BY company_id;
        """

        return sql

    def margin_by_cabang(self, where):
        sql = f"""
                SELECT C.company_name, D.cabang_name, A.order_type, A.sales, A.hpp, A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent FROM (
        SELECT  b.order_type, SUM(C.harga_total) as sales, SUM(C.harga_total_hpp) as hpp, b.cabang_id, b.company_id FROM trans_inventory_subsidiary_invoice A
        LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
        LEFT JOIN trans_inventory_subsidiary_sales_order C ON B.id_trans = C.id_trans
        {where}
        GROUP BY b.company_id, b.cabang_id, b.order_type
        ) A
        LEFT JOIN master_company C on A.company_id = C.id_company
        LEFT JOIN master_company_cabang D on A.company_id = D.id_company and A.cabang_id = D.id_cabang
        ORDER BY company_id, cabang_id;
        """

        return sql

    def margin_by_cabang_month(self, where):
        sql = f"""
                        SELECT C.company_name, D.cabang_name, A.year, A.month, A.order_type, A.sales, A.hpp, A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent FROM (
            SELECT b.order_type, SUM(C.harga_total) as sales, SUM(C.harga_total_hpp) as hpp, b.cabang_id, b.company_id, date_part('month', payment_last_updated) as month, date_part('year', payment_last_updated) as year
            FROM trans_inventory_subsidiary_invoice A
            LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
            LEFT JOIN trans_inventory_subsidiary_sales_order C ON B.id_trans = C.id_trans
            {where}
            GROUP BY b.company_id, b.cabang_id, date_part('year', payment_last_updated), date_part('month', payment_last_updated), b.order_type 
            ) A
            LEFT JOIN master_company C on A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company and A.cabang_id = D.id_cabang
            ORDER BY company_id, cabang_id, a.year, a.month;
                """

        return sql

    def margin_by_invoice(self, where):
        sql = f"""
                                                        SELECT A
            .id_trans,
            C.company_name,
            D.cabang_name,
            A.order_type,
            B.nama_produk,
            A.qty,
            CASE
                WHEN A.produk_id = 4 THEN
                qty * 50 ELSE qty 
            END AS qty_in_kg, A.hpp,
            A.sales,
            A.sales - A.hpp as margin, ROUND((A.sales - A.hpp) / A.sales *100,2)  as margin_percent,
            A.payment_last_updated as tanggal_pelunasan
            FROM
            (
            SELECT A.id_trans, B.order_type, B.company_id, B.cabang_id, C.produk_id, A.payment_last_updated, SUM(c.qty) as qty, SUM(c.harga_total_hpp) AS hpp,  SUM(c.harga_total) as sales 
            FROM trans_inventory_subsidiary_invoice A
            LEFT JOIN trans_inventory_subsidiary_sales_order_header B on A.id_trans_sales_order = B.id_trans
            LEFT JOIN trans_inventory_subsidiary_sales_order C on B.id_trans = C.id_trans
            {where}
            GROUP BY A.id_trans, B.company_id, B.cabang_id, produk_id, A.payment_last_updated, B.order_type
            ) A
            LEFT JOIN master_produk B on A.produk_id = B.id_produk
            LEFT JOIN master_company C on A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company and A.cabang_id = D.id_cabang
            ORDER BY payment_last_updated, company_id, cabang_id;
        """

        return sql


@app.get("/api/f_report/c_margin_report/export_margin_report")
async def get_invoice_so(
    tanggal_awal: str = Query(None, alias="tanggal_awal"),
    tanggal_akhir: str = Query(None, alias="tanggal_akhir"),
    company_id: int = Query(None, alias="company_id"),
    is_range: bool = Query(None, alias="is_range"),
):
    ob_data = c_margin_report()
    return await ob_data.export_margin_report(
        tanggal_awal, tanggal_akhir, company_id, is_range
    )
