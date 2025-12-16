import base64
import calendar
from datetime import datetime
import io

from fastapi import Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from library import *
import os
from library.router import app
from library.db import Db
from pydantic import BaseModel
from modules.f_report.create_paid_sales_report import PDF
from openpyxl.styles import PatternFill, Font


class c_paid_sales_report(object):
    def __init__(self):
        self.db = Db()

    async def pdf_sales_report(self, company_id, tanggal):
        if company_id != None and tanggal != None:
            month = datetime.strptime(tanggal, "%Y-%m-%d").month
            year = datetime.strptime(tanggal, "%Y-%m-%d").year

        sql_header = self.get_query_header(company_id, month, year)
        sql_payment_per_company = self.get_query_per_company(company_id, month, year)
        sql_detail = self.get_query_detail(company_id, month, year)

        query_sql_header = await self.db.executeToDict(sql_header)
        query_sql_detail = await self.db.executeToDict(sql_detail)
        query_sql_payment_per_company = await self.db.executeToDict(
            sql_payment_per_company
        )
        print(sql_detail)

        month_name = calendar.month_name[month]

        pdf = PDF(
            detail_sales_data=query_sql_detail,
            resume_sale_data=query_sql_header,
            resume_sales_payment=query_sql_payment_per_company,
            month_name=month_name,
            year=year,
        )
        pdf_buffer = pdf.generate_report()

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=report.pdf"},
        )

    def write_dict_data(self, ws, data, header):
        HEADER_FONT = Font(b=True, color="000000")
        HEADER_FILL = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        if not data:
            return

        keys = list(data[0].keys())

        # Header
        for col, key in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Rows
        for row in data:
            ws.append([row.get(key) for key in keys])

    async def excel_sales_report(self, company_id, tanggal):

        if not company_id or not tanggal:
            raise ValueError("company_id and tanggal are required")

        date_obj = datetime.strptime(tanggal, "%Y-%m-%d")
        month = date_obj.month
        year = date_obj.year

        sql_header = self.get_query_header(company_id, month, year)
        sql_payment = self.get_query_per_company(company_id, month, year)
        sql_detail = self.get_query_detail(company_id, month, year)

        header_data = await self.db.executeToDict(sql_header)
        payment_data = await self.db.executeToDict(sql_payment)
        detail_data = await self.db.executeToDict(sql_detail)

        wb = Workbook()
        header1 = [
            "Nama Produk",
            "Paid Sales Total",
            "Sales QTY",
            "HPP",
            "Sales_total",
            "Produk ID",
            "Company ID",
            "Nama Company",
            "Harga Satuan Penj",
            "Harga Satuan HPP",
            "Margin Total",
            "Margin Percent",
        ]

        header2 = [
            "Nama Cabang",
            "Nama Produk",
            "Total QTY",
            "Total Sales",
            "Total Outstanding",
            "Total Paid",
            "Paid Percentage",
        ]

        header3 = [
            "No Invoice",
            "Nama Customer",
            "Nama Cabang",
            "Harga Total PPN PPH",
            "Biaya Admin",
            "Harga Total HPP",
            "Margin",
            "Percent Margin",
            "Paid Sales",
            "Paid Margin",
        ]

        sheets = {
            "Produk": {"data": header_data, "header": header1},
            "Cabang": {"data": payment_data, "header": header2},
            "Detail": {"data": detail_data, "header": header3},
        }

        wb.remove(wb.active)

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(sheet_name)
            self.write_dict_data(ws, data["data"], data["header"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def get_query_header(self, company_id, month, year):

        filter_header = f"""
            WHERE date_part('month', payment_last_updated) = {month} AND date_part('year', payment_last_updated) = {year} AND company_id = {company_id}
        AND amount_total != amount_total_outstanding
        """
        sql_header = f"""
                    SELECT
            y.nama_produk,
            x.*,
            z.company_name,
            round( sales_total / sales_qty, 2 ) :: FLOAT AS harga_sat_penj,
            round( hpp / sales_qty, 2 ) :: FLOAT AS harga_sat_hpp,
            sales_total - hpp AS margin_total,
            round( ( sales_total - hpp ) / sales_total * 100, 2 ) :: FLOAT margin_percent 
            FROM
            (
                SELECT SUM(aa.amount_total) - sum(aa.amount_total_outstanding) as paid_sales_total,
                SUM ( bb.qty ) sales_qty,
                SUM ( bb.harga_total_hpp ) AS hpp,
                SUM (aa.amount_total) as sales_total,
                bb.produk_id,
                bb.company_id
                FROM
                trans_inventory_subsidiary_invoice aa
                LEFT JOIN trans_inventory_subsidiary_sales_order bb ON aa.id_trans_sales_order = bb.id_trans
                {filter_header}
                GROUP BY bb.produk_id, bb.company_id
            ) x
            LEFT JOIN master_produk y on x.produk_id = y.id_produk
            LEFT JOIN master_company z on x.company_id = z.id_company

            """

        return sql_header

    def get_query_per_company(self, company_id, month, year):
        filter_branch_detail = f"""
                WHERE date_part('month', payment_last_updated) = {month} AND date_part('year', payment_last_updated) = {year} AND company_id = {company_id} 
            """
        sql_payment_per_company = f"""
            SELECT
                D.cabang_name,
                B.nama_produk,
                A.total_qty,
                A.total_sales,
                A.total_outstanding,
                A.total_sales - A.total_outstanding as paid_sales,
                ROUND((A.total_sales - A.total_outstanding) / A.total_sales * 100, 2)  :: FLOAT as paid_percentage
                FROM
                (
                    SELECT
                    B.company_id,
                    B.cabang_id,
                    B.produk_id,
                    SUM ( B.qty ) AS total_qty,
                    SUM ( amount_total ) AS total_sales,
                    SUM ( amount_total_outstanding ) AS total_outstanding 
                    FROM
                    trans_inventory_subsidiary_invoice
                    A LEFT JOIN trans_inventory_subsidiary_sales_order B ON A.id_trans_sales_order = B.id_trans 
                    {filter_branch_detail}
                    GROUP BY
                    B.produk_id,
                    company_id,
                    cabang_id 
                )
                A LEFT JOIN master_produk B ON A.produk_id = B.id_produk
                LEFT JOIN master_company C ON A.company_id = C.id_company
                LEFT JOIN master_company_cabang D ON A.company_id = D.id_company 
                AND A.cabang_id = D.id_cabang
                ORDER BY company_id, cabang_id, produk_id
        """
        return sql_payment_per_company

    def get_query_detail(self, company_id, month, year):
        filter_detail = f"""
                WHERE date_part('month', payment_last_updated) = {month} AND date_part('year', payment_last_updated) = {year} AND ee.company_id = {company_id}
            AND aa.amount_total != aa.amount_total_outstanding
            """

        sql_detail = f"""
            SELECT
            *,
            ROUND( xx.paid_sales * xx.percent_margin / 100, 2 ) :: FLOAT AS paid_margin 
            FROM
            (
                SELECT
                aa.id_trans AS invoice_number,
                bb.nama_customer,
                hh.cabang_name,
                ee.harga_total_ppn_pph,
                ee.biaya_admin,
                ee.harga_total_hpp,
                ee.harga_total - ee.harga_total_hpp AS margin,
                round( ( ( ee.harga_total - ee.harga_total_hpp ) * 100 / ee.harga_total :: FLOAT ) :: NUMERIC, 2 ) AS percent_margin,
                aa.amount_total - aa.amount_total_outstanding AS paid_sales 
                FROM
                trans_inventory_subsidiary_invoice aa
                LEFT JOIN master_customer bb ON aa.customer_id = bb.id_customer
                LEFT JOIN master_jenis_pembayaran dd ON aa.id_pembayaran = dd.id_pembayaran
                LEFT JOIN trans_inventory_subsidiary_sales_order_header ee ON aa.id_trans_sales_order = ee.id_trans
                LEFT JOIN master_company gg ON ee.company_id = gg.id_company
                LEFT JOIN master_company_cabang hh ON ee.cabang_id = hh.id_cabang
                LEFT JOIN ( SELECT id_user, NAME FROM master_user WHERE is_salesman = 't' ) ii ON ee.salesman = ii.id_user
                LEFT JOIN master_provinsi jj ON bb.kode_prov = jj.kode_prov 
                {filter_detail} 
            AND aa.amount_total != aa.amount_total_outstanding 
            ) XX
                                        """
        return sql_detail


@app.get("/api/f_report/c_paid_sales_report/get_pdf_sales_report")
async def get_pdf_report(
    company_id: int = Query(None, alias="company_id"),
    tanggal: str = Query(None, alias="tanggal"),
):
    ob_data = c_paid_sales_report()
    return await ob_data.pdf_sales_report(company_id, tanggal)


@app.get("/api/f_report/c_paid_sales_report/get_excel_sales_report")
async def get_excel_report(
    company_id: int = Query(None, alias="company_id"),
    tanggal: str = Query(None, alias="tanggal"),
):
    ob_data = c_paid_sales_report()
    return await ob_data.excel_sales_report(company_id, tanggal)
