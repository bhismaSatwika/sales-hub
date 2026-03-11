import base64
import io

from fastapi.responses import StreamingResponse

from library import *
import os
from library.router import app
from library.db import Db
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_stock_report(object):
    def __init__(self):
        self.db = Db()

    async def generate_report(self):
        query = f"""
            SELECT
  B.company_name,
  C.cabang_name,
  D.nama_produk,
  qty_submitted as direct_qty_submmited,
  A.qty_received as direct_qty_received,
  qty_transfered as direct_qty_transfered,
  direct_qty_sold,
  qty_submitted + qty_received  - qty_transfered - direct_qty_sold AS direct_qty_left,
  dropship_qty_delivery,
  dropship_qty_received,
  dropship_qty_sold,
  dropship_qty_delivery + dropship_qty_received - dropship_qty_sold as dropship_qty_left
FROM
  (
        SELECT 
            company_id, cabang_id, produk_id,
            SUM( CASE WHEN mutasi_type = 'ST' AND in_out = 'IN' AND company_id = 1 THEN qty ELSE 0 END) qty_submitted,
            SUM( CASE WHEN mutasi_type = 'TP' AND in_out = 'IN' THEN qty ELSE 0 END) qty_received,
            SUM( CASE WHEN mutasi_type = 'TP' AND in_out = 'OUT' THEN qty ELSE 0 END) qty_transfered,
            SUM( CASE WHEN mutasi_type = 'SO' AND in_out = 'OUT' AND tabel_reference != 'trans_inventory_holding_delivery_preparation_header' THEN qty ELSE 0 END) direct_qty_sold,
            SUM( CASE WHEN mutasi_type = 'DS' AND in_out = 'IN' AND company_id !=1 AND tabel_reference = 'trans_inventory_holding_delivery_preparation_header' THEN qty ELSE 0 END) dropship_qty_received,
            SUM( CASE WHEN mutasi_type = 'SO' AND in_out = 'OUT' AND company_id !=1 AND tabel_reference = 'trans_inventory_holding_delivery_preparation_header' THEN qty ELSE 0 END) dropship_qty_sold
        FROM trans_inventory_detail_mutasi 
        GROUP BY company_id, cabang_id, produk_id
        ) A
        LEFT JOIN master_company B on A.company_id = B.id_company
        LEFT JOIN master_company_cabang C on A.company_id = C.id_company and A.cabang_id = C.id_cabang
        LEFT JOIN master_produk D on A.produk_id = D.id_produk
        LEFT JOIN (
          SELECT A.cabang_id, A.company_id, B.qty as dropship_qty_delivery, B.produk_id FROM trans_inventory_holding_delivery_preparation_header A
          LEFT JOIN trans_inventory_holding_delivery_preparation B on A.id_trans = B.id_trans
          WHERE A.status_release = TRUE
        ) E on A.produk_id = E.produk_id AND A.company_id = E.company_id AND A.cabang_id = E.cabang_id
        ORDER BY a.company_id, a.cabang_id
                """

        res = await self.db.executeToDict(query)
        wb = self.generate_excel(res)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def generate_excel(self, result_data):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "Nama Company"
        ws["B1"].value = "Nama Cabang"
        ws["C1"].value = "Nama Produk"
        ws["D1"].value = "Quantity Submitted"
        ws["E1"].value = "Quantity Received"
        ws["F1"].value = "Quantity Trasnfered"
        ws["G1"].value = "Quantity Dropship"
        ws["H1"].value = "Quantity Terjual"
        ws["I1"].value = "Quantity Sekarang"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb


@app.get("/api/f_report/c_stock_report/get_stock_report")
async def test_get():
    ob_data = c_stock_report()
    return await ob_data.generate_report()
