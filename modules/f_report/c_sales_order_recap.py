from datetime import datetime
import io
import json
import mimetypes
from fastapi import HTTPException, Query, Request
from fastapi.responses import FileResponse, StreamingResponse
from config import params
from library import *
import os
from library.router import app
from library.db import Db
from modules.f_report.create_sales_report import PDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_sales_order_recap(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self, orderby, limit, offset, filter, filter_other="", filter_other_conj=""
    ):

        orderby = "aa.updateindb DESC"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        query = f"""SELECT 
                    ROW_NUMBER() OVER (ORDER BY id DESC) AS nomor_urut,
                    aa.id,
                    aa.tanggal,
                    aa.number_report,
                    aa.status_release,
                    aa.tanggal_start,
                    aa.tanggal_end,
                    (CASE 
                        WHEN aa.status_release = true
                        THEN 'Release'
                    ELSE 'Draft'
                    END) as ket_status_release
                FROM trans_sales_recap_header aa"""

        sql = query + str_clause
        sql_2 = sql + str_clause_count

        sql_count = f"""SELECT COUNT(*) 
        FROM ({sql_2})  as subquery"""

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data

    async def get_max_date(
        self,
    ):

        sql = f"""SELECT MAX(tanggal_end) as max_date FROM trans_sales_recap_header"""

        res = await self.db.executeToDict(sql)
        result = res[0]["max_date"]

        return result

    async def get_id_trans_kode(self, kode_trans, tahun, bulan):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        # sql_kode = (
        #     f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        # )

        # kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT 
                            LPAD( CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                            CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                        FROM trans_sales_recap_header 
                        WHERE DATE_PART('year', tanggal) = {tahun} AND DATE_PART('month', tanggal) = {bulan}"""
        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        number_report = (
            "ALL"
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )
        # print(id_trans)

        data_kode = {
            "number_report": number_report,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode

    async def create(self, data):
        tanggal = datetime.today()
        date = datetime.strptime(data["tanggal_start"], "%Y-%m-%d")
        tahun = date.year
        bulan = date.month

        data_kode = await self.get_id_trans_kode("RPT", tahun, bulan)

        data.update(
            {
                "tanggal": tanggal,
                "tahun": tahun,
                "bulan": bulan,
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
                "number_report": data_kode["number_report"],
                "no_urut": data_kode["no_urut"],
            }
        )

        sqlString = self.db.genStrInsertSingleObject(data, "trans_sales_recap_header")

        try:
            # print(sqlString)
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            print(e)
            message = {"status": "error : " + str(e)}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def update(self, data, data_where):
        tanggal = datetime.today()
        date = datetime.strptime(data["tanggal_start"], "%Y-%m-%d")
        tahun = date.year
        bulan = date.month

        data.update(
            {
                "tanggal": tanggal,
                "tahun": tahun,
                "bulan": bulan,
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
            }
        )

        sqlString = self.db.genUpdateObject(
            data, data_where, "trans_sales_recap_header"
        )
        # print(sqlString)
        try:
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def delete(self, data_where):
        sqlString = self.db.genDeleteObject(data_where, "trans_sales_recap_header")
        try:
            await self.db.executeQuery(sqlString)
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(400, ("The error is: ", str(e)))
        return message

    async def release(self, data):
        data_where = data["data_where"]
        tanggal_start = data_where["tanggal_start"]
        tanggal_end = data_where["tanggal_end"]

        sql_update_st_release_recap_header = f"""UPDATE trans_sales_recap_header SET status_release = 'true'
                                    WHERE id = '{data_where['id']}'"""

        sql_update_invoice = f"""UPDATE trans_inventory_subsidiary_invoice 
            SET id_sales_report = '{data_where['number_report']}' 
            WHERE
            id_sales_report IS NULL AND 
            tanggal_invoice BETWEEN '{tanggal_start}' and '{tanggal_end}'"""

        sql_insert_recap_detail = f"""INSERT INTO trans_sales_recap_detail (id_header,invoice_number) 
                               SELECT '{data_where['id']}',id_trans FROM trans_inventory_subsidiary_invoice
                               where id_sales_report = '{data_where['number_report']}'"""

        # trans = await self.db.executeTrans(
        #     [
        #         sql_update_st_release_recap_header,
        #         sql_update_invoice,
        #         sql_insert_recap_detail,
        #     ]
        # )

        # if trans["status"] == False:
        #     print(str(trans["detail"]))
        #     raise HTTPException(400, str(trans["detail"]))

        sql_resume_sale_direct = f"""
        SELECT
            XX.*,
            AA.nama_produk,
            BB.uom_satuan,
            round( sales_total / sales_qty, 2 ) :: FLOAT AS harga_sat_penj,
            round( hpp / sales_qty, 2 ) :: FLOAT AS harga_sat_hpp,
            sales_total - hpp AS margin_total,
            round( ( sales_total - hpp ) / sales_total * 100, 2 ) :: FLOAT margin_percent 
            FROM
            (
                SELECT SUM
                ( cc.harga_total ) sales_total,
                SUM ( cc.qty ) sales_qty,
                SUM ( cc.harga_total_hpp ) AS hpp,
                produk_id 
                FROM
                trans_sales_recap_detail aa
                LEFT JOIN trans_inventory_subsidiary_invoice bb ON aa.invoice_number = bb.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header D on bb.id_trans_sales_order = D.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order cc ON bb.id_trans_sales_order = cc.id_trans 
                WHERE
                aa.id_header = '{data_where['id']}' and d.order_type = 'direct'
                GROUP BY
            produk_id 
            ) xx
        LEFT JOIN master_produk AA on AA.id_produk = XX.produk_id
        LEFT JOIN master_produk_uom_satuan BB on AA.uom_satuan = BB.id_uom_satuan;
            """

        sql_resume_sale_dropship = f"""
        SELECT
            XX.*,
            AA.nama_produk,
            BB.uom_satuan,
            round( sales_total / sales_qty, 2 ) :: FLOAT AS harga_sat_penj,
            round( hpp / sales_qty, 2 ) :: FLOAT AS harga_sat_hpp,
            sales_total - hpp AS margin_total,
            round( ( sales_total - hpp ) / sales_total * 100, 2 ) :: FLOAT margin_percent 
            FROM
            (
                SELECT SUM
                ( cc.harga_total ) sales_total,
                SUM ( cc.qty ) sales_qty,
                SUM ( cc.harga_total_hpp ) AS hpp,
                produk_id 
                FROM
                trans_sales_recap_detail aa
                LEFT JOIN trans_inventory_subsidiary_invoice bb ON aa.invoice_number = bb.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header D on bb.id_trans_sales_order = D.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order cc ON bb.id_trans_sales_order = cc.id_trans 
                WHERE
                aa.id_header = '{data_where['id']}' and d.order_type = 'dropship'
                GROUP BY
            produk_id 
            ) xx
        LEFT JOIN master_produk AA on AA.id_produk = XX.produk_id
        LEFT JOIN master_produk_uom_satuan BB on AA.uom_satuan = BB.id_uom_satuan;
            """

        sql_resume_inventory = f"""SELECT
                            t.produk_id,
                            SUM(t.qty_in - t.qty_out) AS inv_qty,
                            SUM(t.ht_in - t.ht_out) AS total_hpp,
                            q.uom_satuan,
                            p.nama_produk,
                            ROUND(
                                SUM(t.ht_in - t.ht_out)
                                / NULLIF(SUM(t.qty_in - t.qty_out), 0),
                                2
                            ) AS harga_satuan
                        FROM (
                            SELECT
                                produk_id,
                                SUM(qty) FILTER (WHERE in_out = 'IN')  AS qty_in,
                                SUM(qty) FILTER (WHERE in_out = 'OUT') AS qty_out,
                                SUM(harga_total) FILTER (WHERE in_out = 'IN')  AS ht_in,
                                SUM(harga_total) FILTER (WHERE in_out = 'OUT') AS ht_out
                            FROM trans_inventory_detail_mutasi WHERE
                            tanggal <= '{tanggal_end}'
                            GROUP BY produk_id
                        ) t
                        LEFT JOIN master_produk p
                            ON p.id_produk = t.produk_id
                        LEFT JOIN master_produk_uom_satuan q on p.uom_satuan = q.id_uom_satuan
                        GROUP BY
                            t.produk_id,
                            p.nama_produk,
                            q.uom_satuan;
                            """

        sql_detail_sales_direct = f"""	SELECT
        A.produk_id,
            A.invoice_number,
            B.nama_customer, D.cabang_name, C.company_name, E.nama_produk, A.qty, F.uom_satuan, A.harga_satuan, A.harga_total, A.harga_satuan_hpp, A.harga_total_hpp, A.harga_total - A.harga_total_hpp as margin,
            round(((A.harga_total - A.harga_total_hpp)*100/A.harga_total::FLOAT)::NUMERIC,2) as percent_margin
            FROM
            (
                SELECT A.invoice_number,
                C.customer_id,
                C.company_id,
                C.cabang_id,
                D.produk_id,
                D.harga_satuan,
                D.harga_total,
                D.harga_satuan_hpp,
                D.harga_total_hpp,
                D.qty
                FROM
                trans_sales_recap_detail
                A LEFT JOIN trans_inventory_subsidiary_invoice B ON A.invoice_number = B.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header C ON B.id_trans_sales_order = C.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order D ON C.id_trans = D.id_trans 
                WHERE
                id_header = {data_where['id']} and c.order_type = 'direct'
            ) A 
            LEFT JOIN master_customer B on A.customer_id = B.id_customer
            LEFT JOIN master_company C on A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company AND A.cabang_id = D.id_cabang
            LEFT JOIN master_produk E on A.produk_id = E.id_produk
            LEFT JOIN master_produk_uom_satuan F on E.uom_satuan = F.id_uom_satuan
            ORDER BY
            A.produk_id, A.company_id, A.cabang_id"""

        sql_detail_sales_dropship = f"""	SELECT
        A.produk_id,
            A.invoice_number,
            B.nama_customer, D.cabang_name, C.company_name, E.nama_produk, A.qty, F.uom_satuan, A.harga_satuan, A.harga_total, A.harga_satuan_hpp, A.harga_total_hpp, A.harga_total - A.harga_total_hpp as margin,
            round(((A.harga_total - A.harga_total_hpp)*100/A.harga_total::FLOAT)::NUMERIC,2) as percent_margin
            FROM
            (
                SELECT A.invoice_number,
                C.customer_id,
                C.company_id,
                C.cabang_id,
                D.produk_id,
                D.harga_satuan,
                D.harga_total,
                D.harga_satuan_hpp,
                D.harga_total_hpp,
                D.qty
                FROM
                trans_sales_recap_detail
                A LEFT JOIN trans_inventory_subsidiary_invoice B ON A.invoice_number = B.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header C ON B.id_trans_sales_order = C.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order D ON C.id_trans = D.id_trans 
                WHERE
                id_header = {data_where['id']} and c.order_type = 'dropship'
            ) A 
            LEFT JOIN master_customer B on A.customer_id = B.id_customer
            LEFT JOIN master_company C on A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company AND A.cabang_id = D.id_cabang
            LEFT JOIN master_produk E on A.produk_id = E.id_produk
            LEFT JOIN master_produk_uom_satuan F on E.uom_satuan = F.id_uom_satuan
            ORDER BY
            A.produk_id, A.company_id, A.cabang_id"""

        sql_detail_inventory = f"""SELECT
            C.company_name,
            D.cabang_name,
            A.qty_in - qty_out AS qty,
            CASE
                WHEN ( qty_in - qty_out ) = 0 THEN
                0 ELSE ROUND( ( ht_in - ht_out ) / ( qty_in - qty_out ), 2 ) 
            END AS harga_satuan,
            ht_in - ht_out AS harga_total
            FROM
            (
                SELECT
                produk_id,
                company_id,
                cabang_id,
            SUM ( CASE WHEN in_out = 'IN' THEN qty ELSE 0 END ) qty_in,
            SUM ( CASE WHEN in_out = 'OUT' THEN qty ELSE 0 END ) qty_out,
            SUM ( CASE WHEN in_out = 'IN' THEN harga_total ELSE 0 END ) ht_in,
            SUM ( CASE WHEN in_out = 'OUT' THEN harga_total ELSE 0 END ) ht_out 
            FROM
            trans_inventory_detail_mutasi where
            tanggal <= '{tanggal_end}'
            GROUP BY
            produk_id,
            company_id,
            cabang_id 
            ) A
            LEFT JOIN master_company C ON A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company AND D.id_cabang = A.cabang_id
            WHERE qty_in - qty_out != 0
            ORDER BY A.cabang_id, A.produk_id"""

        resume_sale_direct = await self.db.executeToDict(sql_resume_sale_direct)
        resume_inventory = await self.db.executeToDict(sql_resume_inventory)
        detail_sales_direct = await self.db.executeToDict(sql_detail_sales_direct)
        detail_inventory = await self.db.executeToDict(sql_detail_inventory)

        resume_sale_dropship = await self.db.executeToDict(sql_resume_sale_dropship)
        detail_sales_dropship = await self.db.executeToDict(sql_detail_sales_dropship)

        pdf = PDF(
            number_report=data_where["number_report"],
            resume_sale_data_direct=resume_sale_direct,
            resume_inventory_data=resume_inventory,
            detail_sales_data_direct=detail_sales_direct,
            detail_inventory_data=detail_inventory,
            resume_sale_data_dropship=resume_sale_dropship,
            detail_sales_data_dropship=detail_sales_dropship,
        )
        pdf.generate_report()

    async def get_rekap_resume(self, id_header, id_sales_report):

        return await self.stream_file(
            f"files/sales_order_report/{id_sales_report}.pdf",
            f"{id_sales_report}.pdf",
        )

    def get_content_type(self, file_path):
        # Get the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(file_path)
        # If the MIME type cannot be guessed, fallback to 'application/octet-stream'
        return mime_type if mime_type else "application/octet-stream"

    async def stream_file(self, path, filename):
        try:
            content_type = self.get_content_type(path)
            return FileResponse(path, media_type=content_type, filename=filename)
        except Exception as e:
            raise HTTPException(400, "The error is: " + str(e))

    async def get_excel(self, tanggal_end, id_sales_report, id_header):
        sql_sales_resume = f"""
            SELECT
            AA.nama_produk,
            BB.uom_satuan,
            XX.*,
            round( sales_total / sales_qty, 2 ) :: FLOAT AS harga_sat_penj,
            round( hpp / sales_qty, 2 ) :: FLOAT AS harga_sat_hpp,
            sales_total - hpp AS margin_total,
            round( ( sales_total - hpp ) / sales_total * 100, 2 ) :: FLOAT margin_percent 
            FROM
            (
                SELECT SUM
                ( cc.harga_total ) sales_total,
                SUM ( cc.qty ) sales_qty,
                SUM ( cc.harga_total_hpp ) AS hpp,
                produk_id,
                order_type
                FROM
                trans_sales_recap_detail aa
                LEFT JOIN trans_inventory_subsidiary_invoice bb ON aa.invoice_number = bb.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header D on bb.id_trans_sales_order = D.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order cc ON bb.id_trans_sales_order = cc.id_trans 
                WHERE
                aa.id_header = {id_header}
                GROUP BY
            produk_id, order_type
            ) xx
        LEFT JOIN master_produk AA on AA.id_produk = XX.produk_id
        LEFT JOIN master_produk_uom_satuan BB on AA.uom_satuan = BB.id_uom_satuan;

        """

        sql_inventory_resume = f"""SELECT
                            p.nama_produk,
                            q.uom_satuan,
                            t.produk_id,
                            SUM(t.qty_in - t.qty_out) AS inv_qty,
                            SUM(t.ht_in - t.ht_out) AS total_hpp,
                            ROUND(
                                SUM(t.ht_in - t.ht_out)
                                / NULLIF(SUM(t.qty_in - t.qty_out), 0),
                                2
                            ) AS harga_satuan
                        FROM (
                            SELECT
                                produk_id,
                                SUM(qty) FILTER (WHERE in_out = 'IN')  AS qty_in,
                                SUM(qty) FILTER (WHERE in_out = 'OUT') AS qty_out,
                                SUM(harga_total) FILTER (WHERE in_out = 'IN')  AS ht_in,
                                SUM(harga_total) FILTER (WHERE in_out = 'OUT') AS ht_out
                            FROM trans_inventory_detail_mutasi WHERE
                            tanggal <= '{tanggal_end}'
                            GROUP BY produk_id
                        ) t
                        LEFT JOIN master_produk p
                            ON p.id_produk = t.produk_id
                        LEFT JOIN master_produk_uom_satuan q on p.uom_satuan = q.id_uom_satuan
                        GROUP BY
                            t.produk_id,
                            p.nama_produk,
                            q.uom_satuan;"""

        sql_inventory_detail = f"""
                SELECT
                C.company_name,
                D.cabang_name,
                A.qty_in - qty_out AS qty,
                CASE
                    WHEN ( qty_in - qty_out ) = 0 THEN
                    0 ELSE ROUND( ( ht_in - ht_out ) / ( qty_in - qty_out ), 2 ) 
                END AS harga_satuan,
                ht_in - ht_out AS harga_total
                FROM
                (
                    SELECT
                    produk_id,
                    company_id,
                    cabang_id,
                SUM ( CASE WHEN in_out = 'IN' THEN qty ELSE 0 END ) qty_in,
                SUM ( CASE WHEN in_out = 'OUT' THEN qty ELSE 0 END ) qty_out,
                SUM ( CASE WHEN in_out = 'IN' THEN harga_total ELSE 0 END ) ht_in,
                SUM ( CASE WHEN in_out = 'OUT' THEN harga_total ELSE 0 END ) ht_out 
                FROM
                trans_inventory_detail_mutasi where
                tanggal <= '{tanggal_end}'
                GROUP BY
                produk_id,
                company_id,
                cabang_id 
                ) A
                LEFT JOIN master_company C ON A.company_id = C.id_company
                LEFT JOIN master_company_cabang D on A.company_id = D.id_company AND D.id_cabang = A.cabang_id
                WHERE qty_in - qty_out != 0
                ORDER BY A.cabang_id, A.produk_id
    """

        sql_sales_detail = f"""
            SELECT
        
            A.invoice_number,
            B.nama_customer, D.cabang_name, C.company_name, E.nama_produk, A.order_type, A.qty, F.uom_satuan, A.harga_satuan, A.harga_total, A.harga_satuan_hpp, A.harga_total_hpp, A.harga_total - A.harga_total_hpp as margin,
            round(((A.harga_total - A.harga_total_hpp)*100/A.harga_total::FLOAT)::NUMERIC,2) as percent_margin
            FROM
            (
                SELECT A.invoice_number,
                C.customer_id,
                C.company_id,
                C.cabang_id,
                C.order_type,
                D.produk_id,
                D.harga_satuan,
                D.harga_total,
                D.harga_satuan_hpp,
                D.harga_total_hpp,
                D.qty
                FROM
                trans_sales_recap_detail
                A LEFT JOIN trans_inventory_subsidiary_invoice B ON A.invoice_number = B.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order_header C ON B.id_trans_sales_order = C.id_trans
                LEFT JOIN trans_inventory_subsidiary_sales_order D ON C.id_trans = D.id_trans 
                WHERE
                id_header = {id_header}
            ) A 
            LEFT JOIN master_customer B on A.customer_id = B.id_customer
            LEFT JOIN master_company C on A.company_id = C.id_company
            LEFT JOIN master_company_cabang D on A.company_id = D.id_company AND A.cabang_id = D.id_cabang
            LEFT JOIN master_produk E on A.produk_id = E.id_produk
            LEFT JOIN master_produk_uom_satuan F on E.uom_satuan = F.id_uom_satuan
            ORDER BY
            A.produk_id, A.company_id, A.cabang_id

        """

        sales_resume = await self.db.executeToDict(sql_sales_resume)
        inventory_resume = await self.db.executeToDict(sql_inventory_resume)
        inventory_detail = await self.db.executeToDict(sql_inventory_detail)
        sales_detail = await self.db.executeToDict(sql_sales_detail)

        wb = Workbook()
        sheets = {}
        header1 = [
            "Nama Produk",
            "UOM",
            "Sales",
            "Quantity",
            "HPP",
            "Produk ID",
            "Order Type",
            "Harga Satuan Penjualan",
            "Harga Satuan HPP",
            "Margin Total",
            "Margin Percent",
        ]
        sheets["Sales Resume"] = {
            "data": sales_resume,
            "header": header1,
        }

        header2 = [
            "Nama Produk",
            "UOM",
            "Produk ID",
            "Qty Inventory",
            "Total HPP",
            "Harga Satuan",
        ]
        sheets["Inventory Resume"] = {
            "data": inventory_resume,
            "header": header2,
        }

        header3 = [
            "Nama Company",
            "Nama Cabang",
            "Quantity",
            "Hagra Satuan",
            "Harga Total",
        ]
        sheets["Inventory Detail"] = {
            "data": inventory_detail,
            "header": header3,
        }

        header4 = [
            "No Invoice",
            "Nama Customer",
            "Nama Cabang",
            "Nama Company",
            "Nama Produk",
            "Order Type",
            "Quantity",
            "UOM",
            "Harga Satuan",
            "Harga Total",
            "Harga Satuan HPP",
            "Harga Total HPP",
            "Margin",
            "Margin Percent",
        ]
        sheets["Sales Detail"] = {
            "data": sales_detail,
            "header": header4,
        }

        wb.remove(wb.active)

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(sheet_name)
            self.write_dict_data(ws, data["data"], data["header"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def write_dict_data(self, ws, data, header):
        HEADER_FONT = Font(b=True, color="000000")
        HEADER_FILL = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        if not data:
            return

        keys = list(data[0].keys())

        # Header
        for col, key in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Rows
        for row in data:
            ws.append([row.get(key) for key in keys])


@app.get("/api/f_report/c_sales_order_recap/read")
async def read_data(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
):
    # print("the data:", nik, limit, orderby, offset, filter)
    ob_data = c_sales_order_recap()
    return await ob_data.read(orderby, limit, offset, filter)


@app.get("/api/f_report/c_sales_order_recap/get_max_date")
async def get_max_date():
    ob_data = c_sales_order_recap()
    return await ob_data.get_max_date()


@app.post("/api/f_report/c_sales_order_recap/create")
async def create_data(request: Request):
    data = await request.json()
    ob_data = c_sales_order_recap()
    return await ob_data.create(data)


@app.post("/api/f_report/c_sales_order_recap/update")
async def update_data(request: Request):
    data = await request.json()
    ob_data = c_sales_order_recap()
    return await ob_data.update(data["update_data"], data["update_where"])


@app.post("/api/f_report/c_sales_order_recap/delete")
async def delete(request: Request):
    data = await request.json()
    ob_data = c_sales_order_recap()
    return await ob_data.delete(data)


@app.post("/api/f_report/c_sales_order_recap/release")
async def release(request: Request):
    data = await request.json()
    ob_data = c_sales_order_recap()
    return await ob_data.release(data)


@app.get("/api/f_report/c_sales_order_recap/get_rekap_resume")
async def get_rekap_resume(id, number_report: str):
    ob_data = c_sales_order_recap()
    return await ob_data.get_rekap_resume(id, number_report)


@app.get("/api/f_report/c_sales_order_recap/get_rekap_resume_report")
async def get_rekap_resume_report(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_sales_order_recap()
    path_parent = params.loc["file_sales_recap_report"]
    path = path_parent + id_trans + ".pdf"
    return await ob_data.stream_file(path, id_trans)


@app.get("/api/f_report/c_sales_order_recap/get_excel")
async def get_excel(
    tanggal_end: str = Query(None, alias="tanggal_end"),
    number_report: str = Query(None, alias="number_report"),
    id_header: int = Query(None, alias="id_header"),
):
    ob_data = c_sales_order_recap()
    return await ob_data.get_excel(tanggal_end, number_report, id_header)
