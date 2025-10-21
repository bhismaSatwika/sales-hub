from datetime import datetime
import io
import json
import mimetypes
from typing import List, Optional
from fastapi import HTTPException, Query, Request, Form, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from config import params
from library.router import app
from library.db import Db
from library import *
import os
from modules import f_master
from modules import f_trans
import asyncio
from modules.f_trans.sales_order_create_pdf import PDF
from modules.f_trans.delivery_order_create_pdf import PDF as PDF_DO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_subsidiary_inventory_sales_order(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id=None,
        cabang_id=None,
        filter_other="",
        filter_other_conj="",
    ):

        if company_id != None and cabang_id != None:
            filter_other = (
                f" zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}'"
            )
            filter_other_conj = f" and "

            if company_id == 1:
                filter_other = f""
                filter_other_conj = f""

            if company_id == 2 and cabang_id == 11:
                filter_other = f" zz.company_id = '{company_id}'"

        else:
            filter_other = f""
            filter_other_conj = f""

        if orderby == None or orderby == "":
            orderby = "zz.updateindb DESC"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""SELECT * FROM (
                    SELECT 
                        aa.id_trans,
                        hh.id_trans as id_invoice,
                        bb.id_produk as produk_id,
                        bb.nama_produk||'('||dd.uom_satuan||')' as nama_produk,
                        cc.id_kategori as kategori_id,
                        cc.kategori,
                        dd.id_uom_satuan,
                        dd.uom_satuan,
                        ee.id_company as company_id,
                        ee.company_name,
                        ff.id_cabang as cabang_id,
                        ff.cabang_name,
                        aa.qty,
                        aa.harga_satuan,
                        aa.harga_total,
                        (CASE 
                            WHEN aa.status_release = true
                            THEN 'release'
                            ELSE 'draft'
                        END) as ket_status_release,
                        aa.status_release,
                        aa.tanggal,
                        aa.file_upload,
						gg.id_customer as customer_id,
						gg.nama_customer,
						aa.ppn_percent,
						aa.ppn_value,
						aa.pph_22_percent,
						aa.pph_22_value,
						aa.harga_total_ppn_pph,
                        aa.no_urut,
                        aa.updateindb,
                        gg.account_va,
                        gg.account_bank_name,
                        aa.harga_satuan_hpp,
                        aa.harga_total_hpp,
                        aa.flag_sales_report,
                        hh.md5_file,
                        bb.ppn,
                        bb.pph22,
                        aa.id_pembayaran,
                        ii.pembayaran,
                        aa.salesman,
                        jj.username as nik,
					    jj.name as nama_sales,
                        aa.biaya_admin
                    FROM trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                    LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                    LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
					LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                    LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                    LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                    LEFT JOIN (select * from master_user where is_salesman = 't') jj ON aa.salesman = jj.id_user
                    )zz"""
            + str_clause
        )

        sql_count = (
            f"""SELECT count(*) count FROM (
                    SELECT 
                        aa.id_trans,
                        hh.id_trans as id_invoice,
                        bb.id_produk as produk_id,
                        bb.nama_produk||'('||dd.uom_satuan||')' as nama_produk,
                        cc.id_kategori as kategori_id,
                        cc.kategori,
                        dd.id_uom_satuan,
                        dd.uom_satuan,
                        ee.id_company as company_id,
                        ee.company_name,
                        ff.id_cabang as cabang_id,
                        ff.cabang_name,
                        aa.qty,
                        aa.harga_satuan,
                        aa.harga_total,
                        (CASE 
                            WHEN aa.status_release = true
                            THEN 'release'
                            ELSE 'draft'
                        END) as ket_status_release,
                        aa.status_release,
                        aa.tanggal,
                        aa.file_upload,
						gg.id_customer as customer_id,
						gg.nama_customer,
						aa.ppn_percent,
						aa.ppn_value,
						aa.pph_22_percent,
						aa.pph_22_value,
						aa.harga_total_ppn_pph,
                        aa.no_urut,
                        aa.updateindb,
                        gg.account_va,
                        gg.account_bank_name,
                        aa.harga_satuan_hpp,
                        aa.harga_total_hpp,
                        aa.flag_sales_report,
                        bb.ppn,
                        bb.pph22,
                        aa.id_pembayaran,
                        ii.pembayaran,
                        aa.salesman,
                        jj.username as nik,
					    jj.name as nama_sales,
                        aa.biaya_admin
                    FROM trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                    LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                    LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
					LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                    LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                    LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                    LEFT JOIN (select * from master_user where is_salesman = 't') jj ON aa.salesman = jj.id_user
                    ) zz"""
            + str_clause_count
        )

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data

    async def get_id_trans_kode(self, company_id, cabang_id, kode_trans, tahun, bulan):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )
        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT 
                            LPAD( CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                            CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                        FROM trans_inventory_subsidiary_sales_order 
                        WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND DATE_PART('year', tanggal) = {tahun} AND DATE_PART('month', tanggal) = {bulan}"""
        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode

    async def get_id_trans_kode_release(
        self, company_id, cabang_id, kode_trans, tahun, bulan
    ):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )
        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT
                                LPAD( CAST ( COALESCE ( MAX ( aa.no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                                CAST ( COALESCE ( MAX ( aa.no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                            FROM
                                trans_inventory_subsidiary_delivery_order aa
                                LEFT JOIN trans_inventory_subsidiary_sales_order bb ON aa.id_trans_sales_order = bb.id_trans 
                            WHERE
                                bb.company_id = {company_id}
                                AND bb.cabang_id = {cabang_id} 
                                AND DATE_PART( 'year', aa.tanggal_do ) = {tahun} 
                                AND DATE_PART( 'month', aa.tanggal_do ) = {bulan}"""
        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode

    async def get_id_trans_kode_invoice(
        self, company_id, cabang_id, produk_id, kode_trans, tahun, bulan
    ):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )

        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT
                                LPAD( CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                                CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                            FROM
                                trans_inventory_subsidiary_invoice
                            WHERE
                                produk_id = {produk_id} 
                                AND DATE_PART( 'year', tanggal_invoice ) = {tahun} 
                                AND DATE_PART( 'month', tanggal_invoice ) = {bulan}"""

        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            "IDFOOD."
            + str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode

    async def create(self, data, files: List[UploadFile], listFilename: List[str]):

        tanggal = datetime.strptime(data["tanggal"], "%Y-%m-%d")
        # tanggal = '2025-08-04'
        tahun = tanggal.year
        bulan = tanggal.month

        data_kode = await self.get_id_trans_kode(
            data["company_id"], data["cabang_id"], "SO", tahun, bulan
        )

        sql = f"""select * 
        from trans_inventory_detail 
        where company_id = {data["company_id"]} and cabang_id = {data["cabang_id"]} and produk_id = {data["produk_id"]}"""

        detail_ = await self.db.executeToDict(sql)

        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
                "status_release": False,
                "id_trans": data_kode["id_trans"],
                "no_urut": data_kode["no_urut"],
                "harga_satuan_hpp": detail_[0]["harga_satuan"],
                "harga_total_hpp": int(data["qty"]) * int(detail_[0]["harga_satuan"]),
            }
        )

        # print(data)

        if len(files) > 0:
            path_parent = params.loc["file_inventory_sales_order"]
            file_insert_query = []

            for i, v in enumerate(files):

                filename = data_kode["id_trans"] + "_" + v.filename
                path = path_parent + "/" + filename
                print(path)

                content = await v.read()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                file_ = open(path, "ab")
                file_.write(content)
                file_.close()

                file_insert_query.append(
                    f"""INSERT INTO files_upload (id_trans, file_name, files)
                VALUES('{data_kode["id_trans"]}', '{listFilename[i]}', '{filename}');"""
                )

        # print(data)

        sqlString = self.db.genStrInsertSingleObject(
            data, "trans_inventory_subsidiary_sales_order"
        )

        try:
            await self.db.executeQuery(sqlString)

            if len(files) > 0:
                try:
                    for query in file_insert_query:
                        print(query)
                        await self.db.executeQuery(query)
                except Exception as e:
                    raise HTTPException(400, ("The error is: ", str(e)))

            return "success"
        except Exception as e:
            raise HTTPException(400, ("The error is: ", str(e)))

    async def update(self, data, files: List[UploadFile], listFilename: List[str]):

        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
                "status_release": False,
            }
        )

        sqlString = self.db.genUpdateObject(
            data,
            {"id_trans": data["id_trans"]},
            "trans_inventory_subsidiary_sales_order",
        )

        if len(files) > 0:
            path_parent = params.loc["file_inventory_sales_order"]
            file_insert_query = []

            for i, v in enumerate(files):

                filename = data["id_trans"] + "_" + v.filename
                path = path_parent + "/" + filename
                print(path)

                content = await v.read()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                file_ = open(path, "ab")
                file_.write(content)
                file_.close()

                file_insert_query.append(
                    f"""INSERT INTO files_upload (id_trans, file_name, files)
                VALUES('{data["id_trans"]}', '{listFilename[i]}', '{filename}');"""
                )

        # print(sqlString)
        try:
            await self.db.executeQuery(sqlString)

            if len(files) > 0:
                try:
                    for query in file_insert_query:
                        print(query)
                        await self.db.executeQuery(query)
                except Exception as e:
                    raise HTTPException(400, ("The error is: ", str(e)))

            return "success"
        except Exception as e:
            raise HTTPException(400, ("The error is: ", str(e)))

    async def delete(self, data_where):
        sqlString = self.db.genDeleteObject(
            data_where, "trans_inventory_subsidiary_sales_order"
        )

        sqlString2 = self.db.genDeleteObject(data_where, "files_upload")

        get_files = f"""SELECT files FROM files_upload WHERE id_trans = '{data_where["id_trans"]}'"""

        files = await self.db.executeToDict(get_files)
        # print(files)

        try:
            if len(files) > 0:
                for file in files:
                    path = (
                        params.loc["file_inventory_sales_order"] + "/" + file["files"]
                    )
                    print(path)
                    os.remove(path)
                await self.db.executeTrans([sqlString, sqlString2])
            else:
                await self.db.executeTrans([sqlString, sqlString2])
            message = {"status": "success"}
        except Exception as e:
            message = {"status": "error"}
            raise HTTPException(status_code=400, detail=str(e))
        return message

    async def read_files(self, id_trans):
        sql = f""" SELECT file_name, files FROM files_upload where id_trans = '{id_trans}' """

        result = await self.db.executeToDict(sql)

        output_list = [
            {"file_name": item["file_name"], "file": {"name": item["files"]}}
            for item in result
        ]
        return output_list

    def get_content_type(self, file_path):
        # Get the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(file_path)
        # If the MIME type cannot be guessed, fallback to 'application/octet-stream'
        return mime_type if mime_type else "application/octet-stream"

    async def stream_file(self, path, filename):
        try:
            content_type = self.get_content_type(path)
            return FileResponse(path, media_type=content_type, filename=filename)
        except Exception as e:
            raise HTTPException(400, "The error is: " + str(e))

    async def release(self, data):
        data_where_update = data["data_where_update"]
        save_id_trans_mutasi = ""

        try:
            # --------------------------------------- untuk kebutuhan validasi --------------------------------------------
            sql_valid_detail = f"""SELECT count(aa.*) as count, aa.company_id, aa.cabang_id, aa.produk_id, bb.nama_produk,aa.qty
                            FROM trans_inventory_detail aa
                            LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                            WHERE aa.company_id = {data_where_update['company_id']} AND aa.cabang_id = {data_where_update['cabang_id']} AND aa.produk_id = {data_where_update['produk_id']}
                            GROUP BY aa.company_id, aa.cabang_id, aa.produk_id, bb.nama_produk, aa.qty"""

            result_valid_detail = {
                "inv_detail": await self.db.executeToDict(sql_valid_detail)
            }

            sql_qty_sales_order = f"""select qty from trans_inventory_subsidiary_sales_order where id_trans = '{data_where_update['id_trans']}'"""
            result_qty_sales_order = {
                "inv_sales_order": await self.db.executeToDict(sql_qty_sales_order)
            }

            sql_valid_inventory = f"""SELECT
                                            COALESCE(aa.qty_inventori,0) as qty_inventori,
                                            COALESCE(aa.qty_inventori,0) as qty_validasi
                                        FROM (
                                        SELECT
                                            produk_id,
                                            company_id,
                                            cabang_id,
                                            SUM(qty) as qty_inventori
                                        FROM trans_inventory_detail 
                                        WHERE produk_id = {data_where_update['produk_id']} AND company_id = {data_where_update['company_id']} AND cabang_id =  {data_where_update['cabang_id']}
                                        GROUP BY produk_id,company_id,cabang_id
                                        ) aa"""
            result_inv_valid = {
                "inv_valid": await self.db.executeToDict(sql_valid_inventory)
            }

            if len(result_valid_detail["inv_detail"]) < 1:
                message = "Jumlah Stok tidak ada"
                raise HTTPException(message)

            if (
                result_valid_detail["inv_detail"][0]["count"] != 0
                and result_valid_detail["inv_detail"][0]["count"] != None
                and result_qty_sales_order["inv_sales_order"][0]["qty"]
                <= result_inv_valid["inv_valid"][0]["qty_validasi"]
            ):

                # if True:

                try:
                    sql_sales_order = f"""SELECT
                                        id_trans,
                                        produk_id,
                                        company_id,
                                        cabang_id,
                                        qty,
                                        harga_satuan,
                                        harga_total,
                                        tanggal,
                                        harga_satuan_hpp,
                                        harga_total_hpp
                                    FROM trans_inventory_subsidiary_sales_order
                                    WHERE company_id = {data_where_update['company_id']} AND cabang_id = {data_where_update['cabang_id']} AND produk_id = {data_where_update['produk_id']} and id_trans = '{data_where_update['id_trans']}'"""

                    result_sales_order = {
                        "inv_so": await self.db.executeToDict(sql_sales_order)
                    }

                    data_inv_mutasi = {
                        "produk_id": result_sales_order["inv_so"][0]["produk_id"],
                        "company_id": result_sales_order["inv_so"][0]["company_id"],
                        "cabang_id": result_sales_order["inv_so"][0]["cabang_id"],
                        "qty": result_sales_order["inv_so"][0]["qty"],
                        "harga_satuan": result_sales_order["inv_so"][0][
                            "harga_satuan_hpp"
                        ],
                        "harga_total": result_sales_order["inv_so"][0][
                            "harga_total_hpp"
                        ],
                        "updateindb": datetime.today(),
                        "userupdate": auth.AuthAction.get_data_params("username"),
                        "in_out": "OUT",
                        "mutasi_type": "SO",
                        "id_references": result_sales_order["inv_so"][0]["id_trans"],
                        "tabel_reference": "trans_inventory_subsidiary_sales_order",
                        "tanggal": result_sales_order["inv_so"][0]["tanggal"],
                    }

                    sql_insert_inv_mutasi = (
                        self.db.genStrInsertSingleObject(
                            data_inv_mutasi, "trans_inventory_detail_mutasi"
                        )
                        + " RETURNING id_trans"
                    )

                    # await self.db.executeQuery(sql_insert_inv_mutasi)
                    save = await self.db.executeQueryWithReturn(sql_insert_inv_mutasi)

                    sql_detail_mutasi = f"""SELECT produk_id,
                                        company_id,
                                        cabang_id,
                                        qty,
                                        ROUND(harga_total/qty,0) as harga_satuan,
                                        harga_total 
                                    FROM (
                                    SELECT
                                        produk_id,
                                        company_id,
                                        cabang_id,
                                        SUM(case when in_out ='IN' then qty else 0 end)-SUM(case when in_out ='OUT' then qty else 0 end) qty,
                                        SUM(case when in_out ='IN' then harga_total else 0 end)-SUM(case when in_out ='OUT' then harga_total else 0 end) harga_total
                                    FROM
                                        trans_inventory_detail_mutasi
                                    WHERE company_id = {data_where_update['company_id']} and cabang_id = {data_where_update['cabang_id']} and produk_id = {data_where_update['produk_id']} 
                                        GROUP BY
                                        produk_id,
                                        company_id,
                                        cabang_id
                                        ) aa"""

                    data_detai_mutasi = {
                        "inv_detail_mutasi": await self.db.executeToDict(
                            sql_detail_mutasi
                        )
                    }

                    result_detai_mutasi = data_detai_mutasi
                    save_id_trans_mutasi = save[0]["id_trans"]

                except Exception as e1:

                    where_del = {
                        "id_trans": save_id_trans_mutasi,
                    }

                    sql_del_mutasi_in = self.db.genDeleteObject(
                        where_del, "trans_inventory_detail_mutasi"
                    )

                    try:
                        await self.db.executeQuery(sql_del_mutasi_in)
                    except Exception as e2:
                        print(str(e2))
                        raise HTTPException(
                            400, ("error ketika delete di mutasi: ", str(e2))
                        )

                    print(e1)
                    message = {"status": False, "msg": "error cek query"}
                    raise HTTPException(400, str(e1))

                # await self.create_pdf_do(data_where_update["id_trans"])

                sql_update_status_release_inv_sales_order = f"""UPDATE trans_inventory_subsidiary_sales_order SET status_release = 'true'
                WHERE id_trans = '{data_where_update['id_trans']}'"""

                # ----------------------------------- Perhitungan stock inventori detail --------------------------------------------------------------#
                sql_delete_inv_detail = f"""DELETE FROM trans_inventory_detail 
                                    WHERE company_id = {data_where_update['company_id']} AND cabang_id = {data_where_update['cabang_id']} AND produk_id = {data_where_update['produk_id']}"""

                data_inv_detail_out = {
                    "produk_id": result_detai_mutasi["inv_detail_mutasi"][0][
                        "produk_id"
                    ],
                    "company_id": result_detai_mutasi["inv_detail_mutasi"][0][
                        "company_id"
                    ],
                    "cabang_id": result_detai_mutasi["inv_detail_mutasi"][0][
                        "cabang_id"
                    ],
                    "qty": int(result_detai_mutasi["inv_detail_mutasi"][0]["qty"]),
                    "harga_satuan": int(
                        result_detai_mutasi["inv_detail_mutasi"][0]["harga_satuan"]
                    ),
                    "harga_total": int(
                        result_detai_mutasi["inv_detail_mutasi"][0]["harga_total"]
                    ),
                    "updateindb": datetime.today(),
                    "userupdate": auth.AuthAction.get_data_params("username"),
                }

                sql_insert_inv_detail_out = self.db.genStrInsertSingleObject(
                    data_inv_detail_out, "trans_inventory_detail"
                )

                # ------------------------------------ INSERT KE tabel Delivery Order ------------------------------------------------------------------#
                tanggal = datetime.today()
                tahun = tanggal.year
                bulan = tanggal.month

                data_kode_do = await self.get_id_trans_kode_release(
                    data_where_update["company_id"],
                    data_where_update["cabang_id"],
                    "DO",
                    tahun,
                    bulan,
                )

                sql_inv_sales_order = f"""SELECT * FROM trans_inventory_subsidiary_sales_order WHERE id_trans = '{data_where_update['id_trans']}'"""
                result_inv_sales_order = {
                    "inv_sales_order": await self.db.executeToDict(sql_inv_sales_order)
                }

                data_inv_delivery_order = {
                    "updateindb": datetime.today(),
                    "userupdate": auth.AuthAction.get_data_params("username"),
                    "status_release": False,
                    "tanggal_do": tanggal,
                    "id_trans_sales_order": result_inv_sales_order["inv_sales_order"][
                        0
                    ]["id_trans"],
                    "id_trans": data_kode_do["id_trans"],
                    "no_urut": data_kode_do["no_urut"],
                }

                sql_insert_inv_delivery_order = self.db.genStrInsertSingleObject(
                    data_inv_delivery_order, "trans_inventory_subsidiary_delivery_order"
                )

                # ------------------------------------ INSERT KE tabel Invoice ------------------------------------------------------------------#

                tanggal = datetime.today()
                tahun = tanggal.year
                bulan = tanggal.month

                data_kode_iv = await self.get_id_trans_kode_invoice(
                    data_where_update["company_id"],
                    data_where_update["cabang_id"],
                    result_detai_mutasi["inv_detail_mutasi"][0]["produk_id"],
                    "INV",
                    tahun,
                    bulan,
                )

                data_invoice = {
                    "updateindb": datetime.today(),
                    "userupdate": auth.AuthAction.get_data_params("username"),
                    "status_release": False,
                    "id_trans": data_kode_iv["id_trans"],
                    "tanggal_invoice": tanggal,
                    "id_trans_sales_order": result_inv_sales_order["inv_sales_order"][
                        0
                    ]["id_trans"],
                    "id_trans_delivery_order": data_kode_do["id_trans"],
                    "status_invoice": True,
                    "no_urut": data_kode_iv["no_urut"],
                    "produk_id": result_inv_sales_order["inv_sales_order"][0][
                        "produk_id"
                    ],
                    "amount": result_inv_sales_order["inv_sales_order"][0][
                        "harga_total"
                    ],
                    "amount_ppn": result_inv_sales_order["inv_sales_order"][0][
                        "ppn_value"
                    ],
                    "amount_pph": result_inv_sales_order["inv_sales_order"][0][
                        "pph_22_value"
                    ],
                    "amount_total": result_inv_sales_order["inv_sales_order"][0][
                        "harga_total_ppn_pph"
                    ],
                    "customer_id": result_inv_sales_order["inv_sales_order"][0][
                        "customer_id"
                    ],
                    "qty": result_inv_sales_order["inv_sales_order"][0]["qty"],
                    "salesman": result_inv_sales_order["inv_sales_order"][0][
                        "salesman"
                    ],
                }

                sql_insert_invoice = self.db.genStrInsertSingleObject(
                    data_invoice, "trans_inventory_subsidiary_invoice"
                )

                # eksekusi all transaksi insert, update, delete
                trans = await self.db.executeTrans(
                    [
                        sql_update_status_release_inv_sales_order,
                        sql_delete_inv_detail,
                        sql_insert_inv_detail_out,
                        sql_insert_inv_delivery_order,
                        sql_insert_invoice,
                    ]
                )

                if trans["status"] == False:
                    message = {"status": False, "msg": "Eror. Cek query."}
                    raise HTTPException(400, str(trans["detail"]))

                message = {"status": True, "msg": "success"}

            else:
                # if (
                #     result_qty_sales_order["inv_sales_order"][0]["qty"]
                #     > result_valid_detail["inv_detail"][0]["qty"]
                # ):

                raise HTTPException(
                    "Jumlah stok/kuantiti untuk produk lebih kecil dari jumlah SO. Cek kembali jumlah SO dan stok/kuantiti produk."
                )

        except Exception as e:

            where_del = {
                "id_trans": save_id_trans_mutasi,
            }
            sql_del_mutasi_in = self.db.genDeleteObject(
                where_del, "trans_inventory_detail_mutasi"
            )

            try:
                await self.db.executeQuery(sql_del_mutasi_in)
            except Exception as e2:
                print(str(e2))
                raise HTTPException(400, ("error ketika delete di mutasi: ", str(e2)))

            print(e)
            message = {"status": False, "msg": "Eror. Cek query."}
            raise HTTPException(400, str(e))
        return message

    async def create_pdf_so(self, md5_file):
        sql = f"""SELECT
                    aa.id_trans,
                    bb.id_produk AS produk_id,
                    bb.nama_produk,
                    cc.id_kategori AS kategori_id,
                    cc.kategori,
                    dd.id_uom_satuan,
                    dd.uom_satuan,
                    ee.id_company AS company_id,
                    ee.company_name,
                    ff.id_cabang AS cabang_id,
                    ff.cabang_name,
                    aa.qty,
                    aa.harga_satuan,
                    aa.harga_total,
                    ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'' END ) AS ket_status_release,
                    aa.status_release,
                    aa.tanggal,
                    aa.file_upload,
                    gg.id_customer AS customer_id,
                    gg.nama_customer,
                    aa.ppn_percent,
                    aa.ppn_value,
                    aa.pph_22_percent,
                    aa.pph_22_value,
                    aa.harga_total_ppn_pph,
                    aa.no_urut,
                    aa.updateindb,
                    gg.alamat,
                    gg.no_ktp,
                    gg.no_hp,
                    gg.email,
                    gg.account_va,
                    gg.account_bank_name,
                    hh.id_trans as no_invoice,
                     CASE 
                    WHEN hh.complete_payment THEN
                      'Lunas'
                    ELSE
                      'Belum Lunas'
                    END  as complete_payment,
                    COALESCE(hh.amount_total_outstanding,0) as ato,
                    hh.tanggal_invoice,
                    hh.tanggal_due_date,
                    hh.md5_file,
                    aa.biaya_admin,
                    aa.id_pembayaran,
                    ii.pembayaran,
                    jj.username as nik,
					jj.name as nama_sales
                FROM
                    trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                    LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                    LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang 
                    AND aa.company_id = ff.id_company
                    LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                    LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                    LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                    LEFT JOIN (select * from master_user where is_salesman = 't') jj ON aa.salesman = jj.id_user
                    WHERE hh.md5_file = '{md5_file}'"""

        result = await self.db.executeToDict(sql)

        data = result[0]
        pdf = PDF(data)
        print(data)
        pdf_buffer = pdf.generate_report()
        filenamex = data["id_trans"]

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={filenamex}.pdf"},
        )
        # return data

    async def create_pdf_do(self, id_trans):
        sql = f"""SELECT
                    aa.id_trans,
                    aa.id_trans_sales_order,
                    aa.tanggal_do,
                    bb.customer_id,
                    cc.nama_produk,
                    dd.id_kategori AS kategori_id,
                    dd.kategori,
                    ee.id_uom_satuan,
                    ee.uom_satuan,
                    ff.id_company AS company_id,
                    ff.company_name,
                    gg.id_cabang AS cabang_id,
                    gg.cabang_name,
                    bb.qty,
                    bb.harga_satuan,
                    bb.harga_total,
                    hh.id_customer AS customer_id,
                    hh.nama_customer,
                    hh.alamat,
                    hh.no_ktp,
                    hh.no_hp,
                    hh.email
                FROM
                    trans_inventory_subsidiary_delivery_order aa
                    LEFT JOIN trans_inventory_subsidiary_sales_order bb ON aa.id_trans_sales_order = bb.id_trans
                    LEFT JOIN master_produk cc ON bb.produk_id = cc.id_produk
                    LEFT JOIN master_produk_kategori dd ON cc.kategori_produk = dd.id_kategori
                    LEFT JOIN master_produk_uom_satuan ee ON cc.uom_satuan = ee.id_uom_satuan
                    LEFT JOIN master_company ff ON bb.company_id = ff.id_company
                    LEFT JOIN master_company_cabang gg ON bb.cabang_id = gg.id_cabang 
                    AND bb.company_id = gg.id_company
                    LEFT JOIN master_customer hh ON bb.customer_id = hh.id_customer
                    WHERE bb.id_trans = '{id_trans}'"""

        result = await self.db.executeToDict(sql)

        data = result[0]
        pdf = PDF_DO(data)
        print(data)
        pdf.generate_report()

        return data

    async def read_files(self, id_trans):
        sql = f""" SELECT file_name, files FROM files_upload where id_trans = '{id_trans}' """

        result = await self.db.executeToDict(sql)

        output_list = [
            {"file_name": item["file_name"], "file": {"name": item["files"]}}
            for item in result
        ]
        return output_list

    def get_content_type(self, file_path):
        # Get the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(file_path)
        # If the MIME type cannot be guessed, fallback to 'application/octet-stream'
        return mime_type if mime_type else "application/octet-stream"

    async def stream_file(self, path, filename):
        try:
            content_type = self.get_content_type(path)
            return FileResponse(path, media_type=content_type, filename=filename)
        except Exception as e:
            raise HTTPException(400, "The error is: " + str(e))

    async def delete_file(self, data):
        tbl = "files_upload"
        delete_sql = self.db.genDeleteObject({"files": data["filename"]}, tbl)

        path = str(params.loc["file_inventory_sales_order"]) + "/" + data["filename"]
        try:
            await self.db.executeQuery(delete_sql)

            os.remove(path)
            return "success"
        except Exception as e:
            print(str(e))
            raise HTTPException(400, ("The error is: ", str(e)))

    async def get_invoice_do(self, id_trans):
        sql = f"""select id_trans from trans_inventory_subsidiary_delivery_order where id_trans_sales_order = '{id_trans}'"""
        result = await self.db.executeToDict(sql)
        id_trans_do = result[0]["id_trans"]
        path_parent = params.loc["file_invoice_delivery_order"]
        path = path_parent + id_trans_do + ".pdf"
        return await self.stream_file(path, id_trans_do)

    async def export_sales_order(
        self, tanggal_awal, tanggal_akhir, company_id, cabang_id, is_range
    ):
        company_id = str(company_id)
        cabang_id = str(cabang_id)
        is_range_where = ""
        if is_range:
            is_range_where = " AND tanggal >= '" + tanggal_awal + "'"

        where = (
            "company_id = "
            + company_id
            + " AND cabang_id = "
            + cabang_id
            + " AND tanggal <= '"
            + tanggal_akhir
            + "'"
            + is_range_where
        )

        if int(company_id) == 1 and int(cabang_id) == 1:
            where = "tanggal <= '" + tanggal_akhir + "'" + is_range_where

        elif int(company_id) == 2 and int(cabang_id) == 11:
            where = (
                "company_id = "
                + company_id
                + "AND tanggal <= '"
                + tanggal_akhir
                + "'"
                + is_range_where
            )

        sql = f"""SELECT * FROM (
                    SELECT 
                        aa.id_trans as id_so,
						gg.nama_customer,
                        bb.nama_produk||'('||dd.uom_satuan||')' as nama_produk,
                        aa.tanggal,
                        ee.company_name,
                        ff.cabang_name,
                        aa.qty,
                        aa.harga_satuan,
                        aa.harga_total,
						aa.ppn_percent,
						aa.ppn_value,
						aa.pph_22_percent,
						aa.pph_22_value,
                        aa.biaya_admin,
						aa.harga_total_ppn_pph,                       
                        jj.name as nama_sales,
                        aa.updateindb,
                        aa.company_id,
                        aa.cabang_id
                    FROM trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                    LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                    LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
					LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                    LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                    LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                    LEFT JOIN (select * from master_user where is_salesman = 't') jj ON aa.salesman = jj.id_user
                    )zz 
                    WHERE {where}
                    ORDER BY updateindb DESC
                    """

        result = await self.db.executeToDict(sql)

        wb = self.excel_return(result)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )
        # return result

    def excel_return(self, result_data):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "ID SO"
        ws["B1"].value = "Nama Customer"
        ws["C1"].value = "Nama Produk"
        ws["D1"].value = "Tanggal"
        ws["E1"].value = "Nama Company"
        ws["F1"].value = "Nama Cabang"
        ws["G1"].value = "Quantity"
        ws["H1"].value = "Harga Satuan"
        ws["I1"].value = "Harga Total"
        ws["J1"].value = "PPN %"
        ws["K1"].value = "PPN Value"
        ws["L1"].value = "PPH 22 %"
        ws["M1"].value = "PPH 22 Value"
        ws["N1"].value = "Biaya Admin"
        ws["O1"].value = "Grand Total"
        ws["P1"].value = "SalesMan"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key[:-3]:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb


"""
list your path url at bottom
example /testing url
test from postman :
url/api/c_subsidiary_inventory_sales_order/testing
for post method and other method, check tutorial from 
https://fastapi.tiangolo.com/
"""


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/read")
async def read(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.read(orderby, limit, offset, filter, company_id, cabang_id)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order/create")
async def create(
    produk_id: int = Form(...),
    company_id: int = Form(...),
    cabang_id: int = Form(...),
    qty: int = Form(...),
    harga_satuan: float = Form(...),
    harga_total: float = Form(...),
    tanggal: str = Form(...),
    customer_id: int = Form(...),
    ppn_percent: float = Form(...),
    ppn_value: float = Form(...),
    pph_22_percent: float = Form(...),
    pph_22_value: float = Form(...),
    harga_total_ppn_pph: float = Form(...),
    files: Optional[List[UploadFile]] = File([]),
    filename: Optional[List[str]] = Form(default=[]),
    id_pembayaran: int = Form(...),
    salesman: int = Form(...),
    biaya_admin: float = Form(...),
):
    data = {
        "produk_id": produk_id,
        "company_id": company_id,
        "cabang_id": cabang_id,
        "qty": qty,
        "harga_satuan": harga_satuan,
        "harga_total": harga_total,
        "tanggal": tanggal,
        "customer_id": customer_id,
        "ppn_percent": ppn_percent,
        "ppn_value": ppn_value,
        "pph_22_percent": pph_22_percent,
        "pph_22_value": pph_22_value,
        "harga_total_ppn_pph": harga_total_ppn_pph,
        "id_pembayaran": id_pembayaran,
        "salesman": salesman,
        "biaya_admin": biaya_admin,
    }
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.create(data, files, filename)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order/update")
async def update_data(
    id_trans: str = Form(...),
    produk_id: int = Form(...),
    company_id: int = Form(...),
    cabang_id: int = Form(...),
    qty: int = Form(...),
    harga_satuan: float = Form(...),
    harga_total: float = Form(...),
    tanggal: str = Form(...),
    customer_id: int = Form(...),
    ppn_percent: float = Form(...),
    ppn_value: float = Form(...),
    pph_22_percent: float = Form(...),
    pph_22_value: float = Form(...),
    harga_total_ppn_pph: float = Form(...),
    files: Optional[List[UploadFile]] = File([]),
    filename: Optional[List[str]] = Form(default=[]),
    id_pembayaran: int = Form(...),
    salesman: int = Form(...),
    biaya_admin: float = Form(...),
):
    data = {
        "id_trans": id_trans,
        "produk_id": produk_id,
        "company_id": company_id,
        "cabang_id": cabang_id,
        "qty": qty,
        "harga_satuan": harga_satuan,
        "harga_total": harga_total,
        "tanggal": tanggal,
        "customer_id": customer_id,
        "ppn_percent": ppn_percent,
        "ppn_value": ppn_value,
        "pph_22_percent": pph_22_percent,
        "pph_22_value": pph_22_value,
        "harga_total_ppn_pph": harga_total_ppn_pph,
        "id_pembayaran": id_pembayaran,
        "salesman": salesman,
        "biaya_admin": biaya_admin,
    }
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.update(data, files, filename)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/read_files")
async def get_td_files(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.read_files(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/stream_file")
async def stream_file(filename: str = Query(None, alias="filename")):
    ob_data = c_subsidiary_inventory_sales_order()
    path_parent = params.loc["file_inventory_sales_order"]
    path = path_parent + "/" + filename
    return await ob_data.stream_file(path, filename)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order/delete_file")
async def delete_file(request_: Request):
    data = await request_.json()
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.delete_file(data)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/get_invoice_so")
async def get_invoice_so(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order()
    path_parent = params.loc["file_invoice_sales_order"]
    path = path_parent + id_trans + ".pdf"
    return await ob_data.stream_file(path, id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/get_invoice_do")
async def get_invoice_do(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.get_invoice_do(id_trans)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order/delete")
async def delete(request: Request):
    data = await request.json()
    # data = json.loads(data['param'])
    # data = data['data']
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.delete(data)


# @app.post("/api/f_trans/c_subsidiary_inventory_sales_order/release")
# async def release(request: Request):
#     data = await request.json()
#     ob_data = c_subsidiary_inventory_sales_order()
#     return await ob_data.release(data)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/get_id_trans_kode")
async def get_id_trans_kode(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.get_id_trans_kode(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/get_id_trans_kode")
async def get_id_trans_kode_release(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.get_id_trans_kode_release(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/create_pdf_so")
async def create_pdf_so(id_: str = Query(None, alias="id_")):

    # def replaceForSqlInjection(sqlStr):
    a = ["'", '"']

    for item in a:
        id_ = str(id_).replace(item, "")

    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.create_pdf_so(id_)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/create_pdf_do")
async def create_pdf_do(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.create_pdf_do(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/get_id_trans_kode")
async def get_id_trans_kode(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.get_id_trans_kode(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/read_files")
async def get_td_files(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.read_files(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/stream_file")
async def stream_file(filename: str = Query(None, alias="filename")):
    ob_data = c_subsidiary_inventory_sales_order()
    path_parent = params.loc["file_inventory_sales_order"]
    path = path_parent + "/" + filename
    return await ob_data.stream_file(path, filename)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order/export_sales_order")
async def get_invoice_so(
    tanggal_awal: str = Query(None, alias="tanggal_awal"),
    tanggal_akhir: str = Query(None, alias="tanggal_akhir"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_range: bool = Query(None, alias="is_range"),
):
    ob_data = c_subsidiary_inventory_sales_order()
    return await ob_data.export_sales_order(
        tanggal_awal, tanggal_akhir, company_id, cabang_id, is_range
    )
