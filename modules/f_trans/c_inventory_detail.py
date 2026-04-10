from datetime import datetime
import io
import json
from typing import List
from fastapi import Query, Request, Form, UploadFile, File
from fastapi.responses import StreamingResponse
from library.router import app
from library.db import Db
from library import *
import os
from modules import f_master
from modules import f_trans
import asyncio
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_trans_inventory_detail(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id=None,
        cabang_id=None,
        is_cabang=False,
        is_pusat=False,
        filter_other="",
        filter_other_conj="and",
    ):
        where = "stock_condition = 'good' "

        if company_id != 1 and is_cabang == True:
            filter_other = where + (
                f"AND zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}'"
            )
            filter_other_conj = f" and "

        elif company_id == 1 and is_cabang == True:
            filter_other = where
            filter_other_conj = f"and"

        elif company_id == 1 and is_cabang == False:
            filter_other = where + (
                f"and zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}'"
            )
            filter_other_conj = f" and "

        if company_id != 1 and is_pusat == True and is_cabang == True:
            filter_other = where + f" and zz.company_id = '{company_id}'"
            filter_other_conj = f" and "

        # print(filter_other)
        if orderby == None or orderby == "":
            orderby = "zz.updateindb DESC"
        else:
            orderby = orderby + ", cabang_id asc"

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        # print(str_clause)

        query = """SELECT * FROM (
            SELECT 
                aa.id_trans,
                bb.id_produk as produk_id,
                bb.nama_produk||'('||dd.uom_satuan||')' as nama_produk,
                cc.id_kategori as kategori_id,
                cc.kategori,
                dd.id_uom_satuan,
                dd.uom_satuan,
                ee.id_company as company_id,
                ee.company_name,
                ff.id_cabang as cabang_id,
                ff.cabang_name,
                aa.qty,
                aa.harga_satuan,
                aa.harga_total,
                aa.updateindb,
                bb.ppn,
                bb.pph22,
                aa.stock_condition
            FROM trans_inventory_detail aa
            LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
            LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
            LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
            LEFT JOIN master_company ee ON aa.company_id = ee.id_company
            LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
            LEFT JOIN master_produk_kategori gg ON bb.kategori_produk = gg.id_kategori
            LEFT JOIN master_produk_uom_satuan hh ON bb.uom_satuan = hh.id_uom_satuan
        ) zz """

        # print(sql)

        sql = query + str_clause
        sql_2 = query + str_clause_count
        print(sql)

        sql_count = f"""SELECT COUNT(*) 
        FROM ({sql_2})  as subquery"""

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data

    async def read_product(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id=None,
        cabang_id=None,
        is_cabang=False,
        is_pusat=False,
        filter_other="",
        filter_other_conj="",
    ):

        where = """
        """

        if company_id != 1 and is_cabang == True:
            where = (
                f"where zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}'"
            )

        elif company_id == 1 and is_cabang == False:
            where = (
                f"where zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}'"
            )

        if company_id != 1 and is_pusat == True and is_cabang == True:
            where = f"where zz.company_id = '{company_id}'"

        if orderby == None or orderby == "":
            orderby = "produk_id"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""
        SELECT A.*,
        B.nama_produk,
         c.uom_satuan 
        FROM
        ( SELECT SUM ( qty ) qty, SUM ( harga_total ) AS harga_total, produk_id FROM trans_inventory_detail zz
        {where}
        GROUP BY produk_id )
        A LEFT JOIN master_produk B ON A.produk_id = B.id_produk
        LEFT JOIN master_produk_uom_satuan C ON B.uom_satuan = C.id_uom_satuan
        """
            + str_clause
        )
        # print(sql)

        sql_count = (
            f"""SELECT
        count(*) as count
        FROM 
        ( SELECT SUM ( qty ) qty, SUM ( harga_total ) AS harga_total, produk_id FROM trans_inventory_detail zz
        {where}
        GROUP BY produk_id )
        A LEFT JOIN master_produk B ON A.produk_id = B.id_produk"""
            + str_clause_count
        )

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "total": result_count[0]["count"]}
        return data

    async def read_inventory_card(self, company_id, cabang_id, is_cabang, is_pusat):

        where = """
        """

        if company_id != 1 and is_cabang == True:
            where = f"where company_id = '{company_id}' AND cabang_id = '{cabang_id}'"

        elif company_id == 1 and is_cabang == False:
            where = f"where company_id = '{company_id}' AND cabang_id = '{cabang_id}'"

        if company_id != 1 and is_pusat == True and is_cabang == True:
            where = f"where company_id = '{company_id}'"

        sql_total_branch = f"""
                SELECT SUM
        ( total ) total_branch
        FROM
        ( SELECT COUNT ( DISTINCT cabang_id ) AS total FROM trans_inventory_detail {where} GROUP BY company_id ) x
        """

        sql_total_product = f""" 
        SELECT COUNT
        ( DISTINCT produk_id ) AS total_product
        FROM
        trans_inventory_detail
        {where}
        """

        sql_total_value = f"""
        SELECT SUM
        ( harga_total ) AS total_value 
        FROM
        trans_inventory_detail
        {where}
        """

        print(sql_total_branch)
        print(sql_total_product)
        print(sql_total_value)

        result_total_branch = await self.db.executeToDict(sql_total_branch)
        result_total_product = await self.db.executeToDict(sql_total_product)
        result_total_value = await self.db.executeToDict(sql_total_value)

        data = {
            "total_branch": result_total_branch[0]["total_branch"],
            "total_product": result_total_product[0]["total_product"],
            "total_value": result_total_value[0]["total_value"],
        }
        return data

    async def export_excel(self, company_id, cabang_id, is_cabang, is_pusat):
        where = "WHERE stock_condition = 'good'"

        if company_id != 1 and is_cabang == True:
            filter_other = where + (
                f"AND aa.company_id = '{company_id}' AND aa.cabang_id = '{cabang_id}'"
            )

        elif company_id == 1 and is_cabang == True:
            filter_other = where

        elif company_id == 1 and is_cabang == False:
            filter_other = where + (
                f"and aa.company_id = '{company_id}' AND aa.cabang_id = '{cabang_id}'"
            )

        if company_id != 1 and is_pusat == True and is_cabang == True:
            filter_other = where + f" and aa.company_id = '{company_id}'"

        sql = f"""
        SELECT ee.company_name,
            ff.cabang_name,
            ((bb.nama_produk::text || ' ('::text) || dd.uom_satuan::text) || ')'::text AS nama_produk,
            cc.kategori,
            aa.qty,
            dd.uom_satuan,
            hh.uom_base_convert as qty_convert,
            aa.qty * hh.uom_base_convert AS qty_base,
            hh.uom_base_convert_name,
            aa.harga_satuan,
            aa.harga_total
        FROM trans_inventory_detail aa
            LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
            LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
            LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
            LEFT JOIN master_company ee ON aa.company_id = ee.id_company
            LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
            LEFT JOIN master_produk_kategori gg ON bb.kategori_produk = gg.id_kategori
            LEFT JOIN master_produk_uom_satuan hh ON bb.uom_satuan = hh.id_uom_satuan
            {filter_other}
        ORDER BY aa.produk_id, aa.company_id, aa.cabang_id
            
        """

        result = await self.db.executeToDict(sql)
        print(sql)

        wb = self.excel_return(result)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )

    def excel_return(self, result_data):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "Nama Company"
        ws["B1"].value = "Nama Cabang"
        ws["C1"].value = "Nama Produk"
        ws["D1"].value = "Kategori"
        ws["E1"].value = "Qty"
        ws["F1"].value = "UOM"
        ws["G1"].value = "Convert Qty"
        ws["H1"].value = "Qty Base"
        ws["I1"].value = "UOM Convert"
        ws["J1"].value = "Harga Satuan"
        ws["K1"].value = "Harga Total"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            # print(key, value)
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb


"""
list your path url at bottom
example /testing url
test from postman :
url/api/c_trans_inventory_detail/testing
for post method and other method, check tutorial from 
https://fastapi.tiangolo.com/
"""


@app.get("/api/f_trans/c_trans_inventory_detail/read")
async def read(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_cabang: bool = Query(None, alias="$is_cabang"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_trans_inventory_detail()
    return await ob_data.read(
        orderby, limit, offset, filter, company_id, cabang_id, is_cabang, is_pusat
    )


@app.get("/api/f_trans/c_trans_inventory_detail/export_excel")
async def read(
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_cabang: bool = Query(None, alias="$is_cabang"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_trans_inventory_detail()
    return await ob_data.export_excel(company_id, cabang_id, is_cabang, is_pusat)


@app.get("/api/f_trans/c_trans_inventory_detail/read_product")
async def read_product(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_cabang: bool = Query(None, alias="$is_cabang"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_trans_inventory_detail()
    return await ob_data.read_product(
        orderby, limit, offset, filter, company_id, cabang_id, is_cabang, is_pusat
    )


@app.get("/api/f_trans/c_trans_inventory_detail/read_inventory_card")
async def read_inventory_card(
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
    is_cabang: bool = Query(None, alias="$is_cabang"),
    is_pusat: bool = Query(None, alias="$is_pusat"),
):
    ob_data = c_trans_inventory_detail()
    return await ob_data.read_inventory_card(company_id, cabang_id, is_cabang, is_pusat)
