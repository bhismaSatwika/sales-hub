from datetime import datetime
import io
import json
import mimetypes
from typing import List, Optional
from fastapi import HTTPException, Query, Request, Form, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from config import params
from library.router import app
from library.db import Db
from library import *
import os
from modules import f_master
from modules import f_trans
import asyncio
from modules.f_trans.sales_order_create_pdf import PDF
from modules.f_trans.delivery_order_create_pdf import PDF as PDF_DO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class c_subsidiary_inventory_sales_order_dropship(object):
    def __init__(self):
        self.db = Db()
        self.kendoParse = kendo_parse.KendoParse

    async def read(
        self,
        orderby,
        limit,
        offset,
        filter,
        company_id=None,
        cabang_id=None,
        filter_other="",
        filter_other_conj="",
    ):

        if company_id != None and cabang_id != None:
            filter_other = f" zz.company_id = '{company_id}' AND zz.cabang_id = '{cabang_id}' AND order_type ='dropship'"
            filter_other_conj = f" and "

            if company_id == 1:
                filter_other = f"order_type ='dropship'"
                filter_other_conj = f"AND"

            if company_id == 2 and cabang_id == 11:
                filter_other = (
                    f" zz.company_id = '{company_id}' AND order_type ='dropship'"
                )

        else:
            filter_other = f""
            filter_other_conj = f""

        if orderby == None or orderby == "":
            orderby = "zz.updateindb DESC"
        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""SELECT * FROM (
                    SELECT
                        aa.id,
                        aa.order_type,
                        aa.id_trans,
                        aa.no_urut,
                        aa.company_id,
                        aa.cabang_id,
                        aa.salesman,
                        aa.tanggal,
                        aa.customer_id,
                        aa.id_pembayaran,
                        aa.total_ppn,
                        aa.total_pph,
                        aa.harga_total_hpp,
                        aa.biaya_admin,
                        aa.harga_total_ppn_pph,
                        aa.flag_sales_report,
                        aa.status_release,
                        aa.userupdate,
                        aa.updateindb,
                        ee.company_name,
                        aa.harga_total,
                        ff.cabang_name,
                        ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'draft' END ) AS ket_status_release,
                        gg.nama_customer,
                        gg.account_va,
                        gg.account_bank_name,
                        hh.md5_file,
                        ii.pembayaran,
                        jj.username AS nik,
                        jj.NAME AS nama_sales
                    FROM
                        trans_inventory_subsidiary_sales_order_header aa
                        LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                        LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
                        LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                        LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                        LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                        LEFT JOIN ( SELECT * FROM master_user WHERE is_salesman = 't' ) jj ON aa.salesman = jj.id_user
                    )zz"""
            + str_clause
        )

        sql_count = (
            f"""SELECT count(*) count FROM (
                    SELECT
                        aa.id,
                        aa.order_type,
                        aa.id_trans,
                        aa.no_urut,
                        aa.company_id,
                        aa.cabang_id,
                        aa.salesman,
                        aa.tanggal,
                        aa.customer_id,
                        aa.id_pembayaran,
                        aa.total_ppn,
                        aa.total_pph,
                        aa.harga_total_hpp,
                        aa.biaya_admin,
                        aa.harga_total_ppn_pph,
                        aa.flag_sales_report,
                        aa.status_release,
                        aa.userupdate,
                        aa.updateindb,
                        aa.harga_total,
                        ee.company_name,
                        ff.cabang_name,
                        ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'draft' END ) AS ket_status_release,
                        gg.nama_customer,
                        gg.account_va,
                        gg.account_bank_name,
                        hh.md5_file,
                        ii.pembayaran,
                        jj.username AS nik,
                        jj.NAME AS nama_sales
                    FROM
                        trans_inventory_subsidiary_sales_order_header aa
                        LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                        LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
                        LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                        LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                        LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                        LEFT JOIN ( SELECT * FROM master_user WHERE is_salesman = 't' ) jj ON aa.salesman = jj.id_user
                    ) zz """
            + str_clause_count
        )
        print(sql)

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data
    
    async def read_produk(
        self,
        orderby,
        limit,
        offset,
        filter,
        id_trans=None,
        filter_other="",
        filter_other_conj="",
    ):

        filter_other = f"zz.id_trans = '{id_trans}'"
        filter_other_conj = f" and "
        orderby = "zz.updateindb ASC"

        str_clause = self.kendoParse().parse_query(
            orderby, limit, offset, filter, filter_other, filter_other_conj
        )
        str_clause_count = self.kendoParse().parse_query(
            "", None, None, filter, filter_other, filter_other_conj
        )

        sql = (
            f"""SELECT
                        * 
                    FROM
                        (
                        SELECT
                            aa.id_trans,
                            bb.id_produk AS produk_id,
                            bb.nama_produk || '(' || dd.uom_satuan || ')' AS nama_produk,
                            cc.id_kategori AS kategori_id,
                            cc.kategori,
                            dd.id_uom_satuan,
                            dd.uom_satuan,
                            aa.qty,
                            aa.harga_satuan,
                            aa.harga_total,
                            ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'draft' END ) AS ket_status_release,
                            aa.status_release,
                            aa.tanggal,
                            aa.ppn_percent,
                            aa.ppn_value,
                            aa.pph_22_percent,
                            aa.pph_22_value,
                            aa.harga_total_ppn_pph,
                            aa.no_urut,
                            aa.updateindb,
                            aa.harga_satuan_hpp,
                            aa.harga_total_hpp,
                            aa.flag_sales_report,
                            bb.ppn,
                            bb.pph22,
                            aa.id_pembayaran,
                            aa.salesman,
                            aa.biaya_admin,
                            aa.id_increment
                        FROM
                            trans_inventory_subsidiary_sales_order aa
                            LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                            LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                            LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                        ) zz """
            + str_clause
        )

        sql_count = (
            f"""SELECT
                                count(*) count
                            FROM
                                (
                                SELECT
                                    aa.id_trans,
                                    bb.id_produk AS produk_id,
                                    bb.nama_produk || '(' || dd.uom_satuan || ')' AS nama_produk,
                                    cc.id_kategori AS kategori_id,
                                    cc.kategori,
                                    dd.id_uom_satuan,
                                    dd.uom_satuan,
                                    aa.qty,
                                    aa.harga_satuan,
                                    aa.harga_total,
                                    ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'draft' END ) AS ket_status_release,
                                    aa.status_release,
                                    aa.tanggal,
                                    aa.ppn_percent,
                                    aa.ppn_value,
                                    aa.pph_22_percent,
                                    aa.pph_22_value,
                                    aa.harga_total_ppn_pph,
                                    aa.no_urut,
                                    aa.updateindb,
                                    aa.harga_satuan_hpp,
                                    aa.harga_total_hpp,
                                    aa.flag_sales_report,
                                    bb.ppn,
                                    bb.pph22,
                                    aa.id_pembayaran,
                                    aa.salesman,
                                    aa.biaya_admin,
                                    aa.id_increment
                                FROM
                                    trans_inventory_subsidiary_sales_order aa
                                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                                ) zz """
            + str_clause_count
        )

        result = await self.db.executeToDict(sql)
        result_count = await self.db.executeToDict(sql_count)

        data = {"data": result, "count": result_count[0]["count"]}
        return data
    
    async def get_id_trans_kode(self, company_id, cabang_id, kode_trans, tahun, bulan):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )
        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT 
                            LPAD( CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                            CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                        FROM trans_inventory_subsidiary_sales_order_header
                        WHERE company_id = {company_id} AND cabang_id = {cabang_id} AND DATE_PART('year', tanggal) = {tahun} AND DATE_PART('month', tanggal) = {bulan}"""
        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode

    async def get_id_trans_kode_release(
        self, company_id, cabang_id, kode_trans, tahun, bulan
    ):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )
        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT
                                LPAD( CAST ( COALESCE ( MAX ( aa.no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                                CAST ( COALESCE ( MAX ( aa.no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                            FROM
                                trans_inventory_subsidiary_delivery_order aa
                                LEFT JOIN trans_inventory_subsidiary_sales_order bb ON aa.id_trans_sales_order = bb.id_trans 
                            WHERE
                                bb.company_id = {company_id}
                                AND bb.cabang_id = {cabang_id} 
                                AND DATE_PART( 'year', aa.tanggal_do ) = {tahun} 
                                AND DATE_PART( 'month', aa.tanggal_do ) = {bulan}"""
        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

    async def get_id_trans_kode_invoice(
        self, company_id, cabang_id, produk_id, kode_trans, tahun, bulan
    ):
        # bulan = datetime.now().month
        # tahun = datetime.now().year

        sql_kode = (
            f"""SELECT kode FROM master_company WHERE id_company = {company_id}"""
        )

        kode_company = await self.db.executeToDict(sql_kode)

        sql_no_urut = f"""SELECT
                                LPAD( CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ), 4, '0' ) AS current_no_urut_convert,
                                CAST ( COALESCE ( MAX ( no_urut ), 0 ) + 1 AS VARCHAR ( 32 ) ) AS current_no_urut 
                            FROM
                                trans_inventory_subsidiary_invoice
                            WHERE
                                produk_id = {produk_id} 
                                AND DATE_PART( 'year', tanggal_invoice ) = {tahun} 
                                AND DATE_PART( 'month', tanggal_invoice ) = {bulan}"""

        no_urut = await self.db.executeToDict(sql_no_urut)
        # print(no_urut[0]['current_no_urut_convert'])

        id_trans = (
            "IDFOOD."
            + str(kode_company[0]["kode"])
            + "."
            + str(cabang_id)
            + "."
            + kode_trans
            + "."
            + str(tahun)
            + "."
            + str(str(bulan).zfill(2) + "." + no_urut[0]["current_no_urut_convert"])
        )

        data_kode = {
            "id_trans": id_trans,
            "no_urut": no_urut[0]["current_no_urut_convert"],
        }

        return data_kode
    
    async def create(
        self, data, files: List[UploadFile], listFilename: List[str], product
    ):
        tanggal = datetime.strptime(data["tanggal"], "%Y-%m-%d")
        tahun = tanggal.year
        bulan = tanggal.month

        data_kode = await self.get_id_trans_kode(
            data["company_id"], data["cabang_id"], "SO", tahun, bulan
        )

        # print(data)
        # print("\n\n")
        # print(products)
        # print("\n\n")

        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "id_trans": data_kode["id_trans"],
                "no_urut": data_kode["no_urut"],
            }
        )

        if len(files) > 0:
            path_parent = params.loc["file_inventory_sales_order"]
            file_insert_query = []

            for i, v in enumerate(files):

                filename = data_kode["id_trans"] + "_" + v.filename
                path = path_parent + "/" + filename
                print(path)

                content = await v.read()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                file_ = open(path, "ab")
                file_.write(content)
                file_.close()

                file_insert_query.append(
                    f"""INSERT INTO files_upload (id_trans, file_name, files)
                VALUES('{data_kode["id_trans"]}', '{listFilename[i]}', '{filename}');"""
                )

        sqlHeader = self.db.genStrInsertSingleObject(
            data, "trans_inventory_subsidiary_sales_order_header"
        )

        try:
            products = [
                {
                    **json.loads(p),
                    "id_trans": data_kode["id_trans"],
                    "company_id": data["company_id"],
                    "cabang_id": data["cabang_id"],
                    "userupdate": auth.AuthAction.get_data_params("username"),
                }
                for p in product
            ]

            # print(products)
            # print("\n\n")

            sqlDetail = self.db.genStrInsertArrayObject(
                products, "trans_inventory_subsidiary_sales_order"
            )

            # print(sqlHeader)
            # print(sqlDetail)

            trans = await self.db.executeTrans([sqlHeader, sqlDetail])

            if trans["status"] == False:
                raise HTTPException(400, ("The error is: ", str(e)))

            if len(files) > 0:
                try:
                    for query in file_insert_query:
                        # print(query)
                        await self.db.executeQuery(query)
                except Exception as e:
                    raise HTTPException(400, ("The error is: ", str(e)))

            return "success"
        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

    async def update_produk(self, data):
        produk_update_where = data["product"]["update_where"]
        header_update_where = data["header"]["update_where"]

        produk_update_data = data["product"]["update_data"]
        header_update_data = data["header"]["update_data"]

        sqlProduk = self.db.genUpdateObject(
            produk_update_data,
            produk_update_where,
            "trans_inventory_subsidiary_sales_order",
        )

        sqlHeader = self.db.genUpdateObject(
            header_update_data,
            header_update_where,
            "trans_inventory_subsidiary_sales_order_header",
        )

        try:

            trans = await self.db.executeTrans([sqlProduk, sqlHeader])
            if trans["status"] == False:
                raise HTTPException(400, ("The error is: ", trans["detail"]))

            return "success"

        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))

    async def update(
        self, data, files: List[UploadFile], listFilename: List[str], product: List[str]
    ):
        data.update(
            {
                "userupdate": auth.AuthAction.get_data_params("username"),
                "updateindb": datetime.today(),
                "status_release": False,
            }
        )

        sqlDetail = self.db.genUpdateObject(
            data,
            {"id_trans": data["id_trans"]},
            "trans_inventory_subsidiary_sales_order_header",
        )

        if len(files) > 0:
            path_parent = params.loc["file_inventory_sales_order"]
            file_insert_query = []

            for i, v in enumerate(files):

                filename = data["id_trans"] + "_" + v.filename
                path = path_parent + "/" + filename
                print(path)

                content = await v.read()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                file_ = open(path, "ab")
                file_.write(content)
                file_.close()

                file_insert_query.append(
                    f"""INSERT INTO files_upload (id_trans, file_name, files)
                VALUES('{data["id_trans"]}', '{listFilename[i]}', '{filename}');"""
                )

        products = []

        queries = []

        try:

            if len(product) != 0:

                products = [
                    {
                        **json.loads(p),
                        "id_trans": data["id_trans"],
                        "company_id": data["company_id"],
                        "cabang_id": data["cabang_id"],
                        "userupdate": auth.AuthAction.get_data_params("username"),
                    }
                    for p in product
                ]

                sqlProduct = self.db.genStrInsertArrayObject(
                    products, "trans_inventory_subsidiary_sales_order"
                )
                print(sqlProduct)

                queries.append(str(sqlProduct))

            queries.append(str(sqlDetail))

            trans = await self.db.executeTrans(queries)

            if trans["status"] == False:
                raise HTTPException(400, ("The error is: ", str(e)))

            if len(files) > 0:
                try:
                    for query in file_insert_query:
                        print(query)
                        await self.db.executeQuery(query)
                except Exception as e:
                    print(e)
                    raise HTTPException(400, ("The error is: ", str(e)))

            return "success"
        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))
        
    async def delete_produk(self, data):
        produk_delete_where = data["product"]["update_where"]
        header_update_where = data["header"]["update_where"]
        header_update_data = data["header"]["update_data"]

        sqlProduk = self.db.genDeleteObject(
            produk_delete_where, "trans_inventory_subsidiary_sales_order"
        )

        sqlHeader = self.db.genUpdateObject(
            header_update_data,
            header_update_where,
            "trans_inventory_subsidiary_sales_order_header",
        )

        try:
            trans = await self.db.executeTrans([sqlProduk, sqlHeader])

            if trans["status"] == False:
                raise HTTPException(400, ("The error is: ", str(e)))

            message = {"status": "success"}
        except Exception as e:
            print(e)
            message = {"status": "error"}
            raise HTTPException(status_code=400, detail=str(e))
        return message
    
    async def delete(self, data_where):
        sqlHeader = self.db.genDeleteObject(
            data_where, "trans_inventory_subsidiary_sales_order_header"
        )

        sqlDetail = self.db.genDeleteObject(
            data_where, "trans_inventory_subsidiary_sales_order"
        )

        sqlFile = self.db.genDeleteObject(data_where, "files_upload")

        get_files = f"""SELECT files FROM files_upload WHERE id_trans = '{data_where["id_trans"]}'"""
        files = await self.db.executeToDict(get_files)

        try:
            if len(files) > 0:
                trans = await self.db.executeTrans([sqlHeader, sqlDetail, sqlFile])

                if trans["status"] == False:
                    raise HTTPException(400, ("The error is: ", str(trans["detail"])))

                for file in files:
                    path = (
                        params.loc["file_inventory_sales_order"] + "/" + file["files"]
                    )
                    print(path)
                    os.remove(path)

            else:
                await self.db.executeTrans([sqlHeader, sqlDetail, sqlFile])
            message = {"status": "success"}
        except Exception as e:
            print(e)
            message = {"status": "error"}
            raise HTTPException(status_code=400, detail=str(e))
        return message
        
    async def read_files(self, id_trans):
        sql = f""" SELECT file_name, files FROM files_upload where id_trans = '{id_trans}' """

        result = await self.db.executeToDict(sql)

        output_list = [
            {"file_name": item["file_name"], "file": {"name": item["files"]}}
            for item in result
        ]
        return output_list

    def get_content_type(self, file_path):
        # Get the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(file_path)
        # If the MIME type cannot be guessed, fallback to 'application/octet-stream'
        return mime_type if mime_type else "application/octet-stream"
    
    async def stream_file(self, path, filename):
        try:
            content_type = self.get_content_type(path)
            return FileResponse(path, media_type=content_type, filename=filename)
        except Exception as e:
            raise HTTPException(400, "The error is: " + str(e))

    async def create_pdf_so(self, id_trans):
        sql_header = f"""SELECT
                            aa.id,
                            aa.id_trans,
                            aa.no_urut,
                            aa.company_id,
                            aa.cabang_id,
                            aa.salesman,
                            aa.tanggal,
                            aa.customer_id,
                            aa.id_pembayaran,
                            aa.total_ppn,
                            aa.total_pph,
                            aa.harga_total_hpp,
                            aa.biaya_admin,
                            aa.harga_total_ppn_pph,
                            aa.flag_sales_report,
                            aa.status_release,
                            aa.userupdate,
                            aa.updateindb,
                            aa.harga_total,
                            bb.company_name,
                            cc.cabang_name,
                            dd.name as nama_sales,
                            ee.pembayaran,
                            ( CASE WHEN aa.status_release = TRUE THEN 'release' ELSE'draft' END ) AS ket_status_release,
                            ff.id_trans as no_invoice,
                            ff.tanggal_invoice,
                            gg.nama_customer,
                            gg.alamat,
                            gg.account_va,
                            gg.no_hp,
                            ff.tanggal_due_date,
                            ff.amount_total_outstanding as ato,
                            gg.account_bank_name,
                            ff.md5_file,
                            CASE WHEN ff.complete_payment = TRUE THEN 'Lunas' ELSE 'Belum Lunas' END as complete_payment
                        FROM trans_inventory_subsidiary_sales_order_header aa
                        LEFT JOIN master_company bb ON aa.company_id = bb.id_company
                        LEFT JOIN master_company_cabang cc ON aa.company_id = bb.id_company AND aa.cabang_id = cc.id_cabang
                        LEFT JOIN master_user dd ON aa.salesman = dd.id_user
                        LEFT JOIN master_jenis_pembayaran ee ON aa.id_pembayaran = ee.id_pembayaran
                        LEFT JOIN trans_inventory_subsidiary_invoice ff ON aa.id_trans = ff.id_trans_sales_order
                        LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                        WHERE ff.md5_file = '{id_trans}'"""

        print(sql_header)

        sql_detail = f"""SELECT
                            dd.nama_produk,
                            aa.qty,
                            ee.uom_satuan,
                            aa.harga_satuan,
                            aa.harga_total,
                            aa.pph_22_value,
                            aa.ppn_value,
                            aa.harga_total_ppn_pph
                    FROM trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_company bb ON aa.company_id = bb.id_company
                    LEFT JOIN master_company_cabang cc ON aa.company_id = bb.id_company AND aa.cabang_id = cc.id_cabang
                    LEFT JOIN master_produk dd ON aa.produk_id = dd.id_produk
                    LEFT JOIN master_produk_uom_satuan ee ON dd.uom_satuan = ee.id_uom_satuan
                    LEFT JOIN trans_inventory_subsidiary_invoice ff ON aa.id_trans = ff.id_trans_sales_order
                    WHERE ff.md5_file = '{id_trans}'"""

        # result = await self.db.executeTrans([sql_header,sql_detail])
        result_header = await self.db.executeToDict(sql_header)
        result_detail = await self.db.executeToDict(sql_detail)

        data_header = result_header[0]
        data_detail = result_detail
        pdf = PDF(data_header, data_detail)

        pdf_buffer = pdf.generate_report()
        filenamex = data_header["id_trans"]

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={filenamex}.pdf"},
        )
    
    async def create_pdf_do(self, id_trans):
        sql = f"""SELECT
                    aa.id_trans,
                    aa.id_trans_sales_order,
                    aa.tanggal_do,
                    bb.customer_id,
                    cc.nama_produk,
                    dd.id_kategori AS kategori_id,
                    dd.kategori,
                    ee.id_uom_satuan,
                    ee.uom_satuan,
                    ff.id_company AS company_id,
                    ff.company_name,
                    gg.id_cabang AS cabang_id,
                    gg.cabang_name,
                    bb.qty,
                    bb.harga_satuan,
                    bb.harga_total,
                    hh.id_customer AS customer_id,
                    hh.nama_customer,
                    hh.alamat,
                    hh.no_ktp,
                    hh.no_hp,
                    hh.email
                FROM
                    trans_inventory_subsidiary_delivery_order aa
                    LEFT JOIN trans_inventory_subsidiary_sales_order bb ON aa.id_trans_sales_order = bb.id_trans
                    LEFT JOIN master_produk cc ON bb.produk_id = cc.id_produk
                    LEFT JOIN master_produk_kategori dd ON cc.kategori_produk = dd.id_kategori
                    LEFT JOIN master_produk_uom_satuan ee ON cc.uom_satuan = ee.id_uom_satuan
                    LEFT JOIN master_company ff ON bb.company_id = ff.id_company
                    LEFT JOIN master_company_cabang gg ON bb.cabang_id = gg.id_cabang 
                    AND bb.company_id = gg.id_company
                    LEFT JOIN master_customer hh ON bb.customer_id = hh.id_customer
                    WHERE bb.id_trans = '{id_trans}'"""

        result = await self.db.executeToDict(sql)

        data = result[0]
        pdf = PDF_DO(data)
        print(data)
        pdf.generate_report()

        return data

    async def read_files(self, id_trans):
        sql = f""" SELECT file_name, files FROM files_upload where id_trans = '{id_trans}' """

        result = await self.db.executeToDict(sql)

        output_list = [
            {"file_name": item["file_name"], "file": {"name": item["files"]}}
            for item in result
        ]
        return output_list

    def get_content_type(self, file_path):
        # Get the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(file_path)
        # If the MIME type cannot be guessed, fallback to 'application/octet-stream'
        return mime_type if mime_type else "application/octet-stream"

    async def delete_file(self, data):
        tbl = "files_upload"
        delete_sql = self.db.genDeleteObject({"files": data["filename"]}, tbl)

        path = str(params.loc["file_inventory_sales_order"]) + "/" + data["filename"]
        try:
            await self.db.executeQuery(delete_sql)

            os.remove(path)
            return "success"
        except Exception as e:
            print(str(e))
            raise HTTPException(400, ("The error is: ", str(e)))

    async def get_invoice_do(self, id_trans):
        sql = f"""select id_trans from trans_inventory_subsidiary_delivery_order where id_trans_sales_order = '{id_trans}'"""
        result = await self.db.executeToDict(sql)
        id_trans_do = result[0]["id_trans"]
        path_parent = params.loc["file_invoice_delivery_order"]
        path = path_parent + id_trans_do + ".pdf"
        return await self.stream_file(path, id_trans_do)

    async def export_sales_order(
        self, tanggal_awal, tanggal_akhir, company_id, cabang_id, is_range
    ):
        company_id = str(company_id)
        cabang_id = str(cabang_id)
        is_range_where = ""
        if is_range:
            is_range_where = " AND tanggal >= '" + tanggal_awal + "'"

        where = (
            "status_release = true and company_id = "
            + company_id
            + " AND cabang_id = "
            + cabang_id
            + " AND tanggal <= '"
            + tanggal_akhir
            + "'"
            + is_range_where
        )

        if int(company_id) == 1 and int(cabang_id) == 1:
            where = (
                "status_release = true and tanggal <= '"
                + tanggal_akhir
                + "'"
                + is_range_where
            )

        elif int(company_id) == 2 and int(cabang_id) == 11:
            where = (
                "status_release = true and company_id = "
                + company_id
                + "AND tanggal <= '"
                + tanggal_akhir
                + "'"
                + is_range_where
            )

        sql = f"""SELECT * FROM (
                     SELECT 
                        aa.id_trans as id_so,
                        hh.id_trans as id_invoice,
                        gg.account_va,
						gg.nama_customer,
                        gg.npwp,
                        bb.nama_produk||'('||dd.uom_satuan||')' as nama_produk,
                        aa.tanggal,
                        ee.company_name,
                        ff.cabang_name,
                        aa.qty,
                        aa.harga_satuan,
                        aa.harga_total,
						aa.ppn_percent,
						aa.ppn_value,
						aa.pph_22_percent,
						aa.pph_22_value,
                        aa.biaya_admin,
						aa.harga_total_ppn_pph,
                        ii.pembayaran,                       
                        jj.name as nama_sales,
                        aa.updateindb,
                        aa.company_id,
                        aa.cabang_id,
                        aa.status_release
                    FROM trans_inventory_subsidiary_sales_order aa
                    LEFT JOIN master_produk bb ON aa.produk_id = bb.id_produk
                    LEFT JOIN master_produk_kategori cc ON bb.kategori_produk = cc.id_kategori
                    LEFT JOIN master_produk_uom_satuan dd ON bb.uom_satuan = dd.id_uom_satuan
                    LEFT JOIN master_company ee ON aa.company_id = ee.id_company
                    LEFT JOIN master_company_cabang ff ON aa.cabang_id = ff.id_cabang AND aa.company_id = ff.id_company
					LEFT JOIN master_customer gg ON aa.customer_id = gg.id_customer
                    LEFT JOIN trans_inventory_subsidiary_invoice hh ON aa.id_trans = hh.id_trans_sales_order
                    LEFT JOIN master_jenis_pembayaran ii ON aa.id_pembayaran = ii.id_pembayaran
                    LEFT JOIN (select * from master_user where is_salesman = 't') jj ON aa.salesman = jj.id_user
                    )zz 
                    WHERE {where}
                    ORDER BY updateindb DESC
                    """

        result = await self.db.executeToDict(sql)
        print(sql)

        wb = self.excel_return(result)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=example.xlsx"},
        )
        # return result

    def excel_return(self, result_data):

        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "ID SO"
        ws["B1"].value = "ID Invoice"
        ws["C1"].value = "Account VA"
        ws["D1"].value = "Nama Customer"
        ws["E1"].value = "NPWP"
        ws["F1"].value = "Nama Produk"
        ws["G1"].value = "Tanggal"
        ws["H1"].value = "Nama Company"
        ws["I1"].value = "Nama Cabang"
        ws["J1"].value = "Quantity"
        ws["K1"].value = "Harga Satuan"
        ws["L1"].value = "Harga Total"
        ws["M1"].value = "PPN %"
        ws["N1"].value = "PPN Value"
        ws["O1"].value = "PPH 22 %"
        ws["P1"].value = "PPH 22 Value"
        ws["Q1"].value = "Biaya Admin"
        ws["R1"].value = "Grand Total"
        ws["S1"].value = "Pembayaran"
        ws["T1"].value = "SalesMan"

        if len(result_data) > 0:
            data_key = []
            i = 0

        x = 0
        for key, value in result_data[0].items():
            print(key, value)
            data_key.append(key)
            ws.cell(1, x + 1).font = Font(b=True, color="000000")
            ws.cell(1, x + 1).fill = PatternFill(
                start_color="ffff00", end_color="ffff00", fill_type="solid"
            )
            x = x + 1

        for data in result_data:
            data_export = []
            for key in data_key[:-4]:
                data_export.append(data[key])
            ws.append(data_export)
            i = i + 1

        return wb

    async def request_approve(self, data):  

        queries = []
        if data["status_release"] == True:
            sql = f"""
            UPDATE trans_approval_detail SET active = false WHERE header_id = '{data["id_trans"]}'
        """
            sql_update = f"""UPDATE trans_inventory_subsidiary_sales_order_header SET approval_status = 1 WHERE id_trans = '{data["id_trans"]}'"""
            queries.append(sql)
            queries.append(sql_update)

        else:
            sql_update_status_release = f"""UPDATE trans_inventory_subsidiary_sales_order_header SET status_release = 'TRUE'
        WHERE id_trans = '{data["id_trans"]}'"""
            queries.append(sql_update_status_release)

        now = datetime.now()
        id = int(now.timestamp()*1000)
        approval_trans = "DROPSHP.APPR."+str(id)

        # insert ke tabel detail_approval
        sql_detail_approval = f"""INSERT INTO trans_approval_detail (header_id,master_approval_id,order_approve,approval_status,approval_type,approval_trans) 
        SELECT 
            '{data["id_trans"]}' AS header_id,
            id AS master_approval_id,
            approval_order AS order_approve,
            (CASE 
                WHEN approval_order = 1
                    THEN 1
                ELSE 2
            END) AS approval_status,
            approval_type AS approval_type,
            '{approval_trans}' as approval_trans
        FROM master_approval 
        WHERE approval_company_id = {data["company_id"]} AND approval_cabang_id = {data["cabang_id"]}"""
        queries.append(sql_detail_approval)

        # sql_insert_detail_approval = self.db.genStrInsertSingleObject(
        #     sql_detail_approval, "trans_inventory_subsidiary_retur_detail"
        # )

        try:
            trans = await self.db.executeTrans(queries)
            if trans["status"] == False:
                raise HTTPException(
                    400, ("error ketika request approval: ", trans["message"])
                )
            return "success"
        except Exception as e:
            print(e)
            raise HTTPException(400, ("The error is: ", str(e)))
        

"""
list your path url at bottom
example /testing url
test from postman :
url/api/c_subsidiary_inventory_sales_order_dropship/testing
for post method and other method, check tutorial from 
https://fastapi.tiangolo.com/
"""

@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/read")
async def read(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    company_id: int = Query(None, alias="$company_id"),
    cabang_id: int = Query(None, alias="$cabang_id"),
):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.read(orderby, limit, offset, filter, company_id, cabang_id)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/read_produk")
async def read_produk(
    limit: int = Query(None, alias="$top"),
    orderby: str = Query(None, alias="$orderby"),
    offset: int = Query(None, alias="$skip"),
    filter: str = Query(None, alias="$filter"),
    id_trans: str = Query(None, alias="id_trans"),
):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.read_produk(orderby, limit, offset, filter, id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_id_trans_kode")
async def get_id_trans_kode(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_id_trans_kode(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_id_trans_kode")
async def get_id_trans_kode_release(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_id_trans_kode_release(
        company_id, cabang_id, kode_trans, tahun, bulan
    )

@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/create")
async def create(
    company_id: int = Form(...),
    cabang_id: int = Form(...),
    tanggal: str = Form(...),
    customer_id: int = Form(...),
    harga_total_ppn_pph: float = Form(...),
    total_ppn: float = Form(...),
    total_pph: float = Form(...),
    files: Optional[List[UploadFile]] = File([]),
    filename: Optional[List[str]] = Form(default=[]),
    id_pembayaran: int = Form(...),
    salesman: int = Form(...),
    biaya_admin: float = Form(...),
    product: List[str] = Form(...),
    harga_total_hpp: float = Form(...),
    harga_total: float = Form(...),
):
    data = {
        "company_id": company_id,
        "cabang_id": cabang_id,
        "tanggal": tanggal,
        "customer_id": customer_id,
        "harga_total_ppn_pph": harga_total_ppn_pph,
        "total_ppn": total_ppn,
        "total_pph": total_pph,
        "id_pembayaran": id_pembayaran,
        "salesman": salesman,
        "biaya_admin": biaya_admin,
        "harga_total_hpp": harga_total_hpp,
        "harga_total": harga_total,
    }

    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.create(data, files, filename, product)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/update_produk")
async def update_produk(request: Request):
    data = await request.json()
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.update_produk(data)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/update")
async def update(
    id_trans: str = Form(...),
    company_id: int = Form(...),
    cabang_id: int = Form(...),
    salesman: int = Form(...),
    tanggal: str = Form(...),
    customer_id: int = Form(...),
    id_pembayaran: int = Form(...),
    total_ppn: float = Form(...),
    total_pph: float = Form(...),
    harga_total_hpp: float = Form(...),
    biaya_admin: float = Form(...),
    harga_total_ppn_pph: float = Form(...),
    harga_total: float = Form(...),
    files: Optional[List[UploadFile]] = File([]),
    filename: Optional[List[str]] = Form(default=[]),
    product: Optional[List[str]] = Form(default=[]),
):
    data = {
        "id_trans": id_trans,
        "company_id": company_id,
        "cabang_id": cabang_id,
        "salesman": salesman,
        "tanggal": tanggal,
        "customer_id": customer_id,
        "id_pembayaran": id_pembayaran,
        "total_ppn": total_ppn,
        "total_pph": total_pph,
        "harga_total_hpp": harga_total_hpp,
        "biaya_admin": biaya_admin,
        "harga_total_ppn_pph": harga_total_ppn_pph,
        "harga_total": harga_total,
    }
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.update(data, files, filename, product)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/read_files")
async def get_td_files(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.read_files(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/stream_file")
async def stream_file(filename: str = Query(None, alias="filename")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    path_parent = params.loc["file_inventory_sales_order"]
    path = path_parent + "/" + filename
    return await ob_data.stream_file(path, filename)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/delete_file")
async def delete_file(request_: Request):
    data = await request_.json()
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.delete_file(data)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_invoice_so")
async def get_invoice_so(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    path_parent = params.loc["file_invoice_sales_order"]
    path = path_parent + id_trans + ".pdf"
    return await ob_data.stream_file(path, id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_invoice_do")
async def get_invoice_do(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_invoice_do(id_trans)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/delete_produk")
async def delete_produk(request: Request):
    data = await request.json()
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.delete_produk(data)


@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/delete")
async def delete(request: Request):
    data = await request.json()
    # data = json.loads(data['param'])
    # data = data['data']
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.delete(data)


# @app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/release")
# async def release(request: Request):
#     data = await request.json()
#     ob_data = c_subsidiary_inventory_sales_order_dropship()
#     return await ob_data.release(data)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_id_trans_kode")
async def get_id_trans_kode(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_id_trans_kode(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_id_trans_kode")
async def get_id_trans_kode_release(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_id_trans_kode_release(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/create_pdf_so")
async def create_pdf_so(id_: str = Query(None, alias="id_")):

    # def replaceForSqlInjection(sqlStr):
    a = ["'", '"']

    for item in a:
        id_ = str(id_).replace(item, "")

    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.create_pdf_so(id_)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/create_pdf_do")
async def create_pdf_do(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.create_pdf_do(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/get_id_trans_kode")
async def get_id_trans_kode(company_id, cabang_id, kode_trans, tahun, bulan):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.get_id_trans_kode(
        company_id, cabang_id, kode_trans, tahun, bulan
    )


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/read_files")
async def get_td_files(id_trans: str = Query(None, alias="id_trans")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.read_files(id_trans)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/stream_file")
async def stream_file(filename: str = Query(None, alias="filename")):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    path_parent = params.loc["file_inventory_sales_order"]
    path = path_parent + "/" + filename
    return await ob_data.stream_file(path, filename)


@app.get("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/export_sales_order")
async def get_invoice_so(
    tanggal_awal: str = Query(None, alias="tanggal_awal"),
    tanggal_akhir: str = Query(None, alias="tanggal_akhir"),
    company_id: int = Query(None, alias="company_id"),
    cabang_id: int = Query(None, alias="cabang_id"),
    is_range: bool = Query(None, alias="is_range"),
):
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    return await ob_data.export_sales_order(
        tanggal_awal, tanggal_akhir, company_id, cabang_id, is_range
    )

@app.post("/api/f_trans/c_subsidiary_inventory_sales_order_dropship/request_approve")
async def request_approve(request: Request):
    data = await request.json()
    ob_data = c_subsidiary_inventory_sales_order_dropship()
    data = data["data_where_update"]
    print(data)
    return await ob_data.request_approve(data)